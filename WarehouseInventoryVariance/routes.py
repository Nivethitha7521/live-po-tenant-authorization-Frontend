from fastapi import APIRouter, HTTPException, Query, Body,Request
from typing import List, Optional, Dict
from datetime import datetime, timedelta
from dependencies.auth import validate_token
from middlewares.permission_middleware import check_permission
from fastapi import Depends
import asyncio
import math
from openpyxl.styles import Font, PatternFill
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
import io
from datetime import timedelta, timezone

from OutletInventory.funtions import ist_now
from db.collections import (
    approvedstocks_rm_collection,
    closingstocks_rm_collection,
    grn_collection,
    inventory_stock_collection,
    purchaseitem_collection,
    storeDispatch_collection,
    warehouse_collection,
    purchase_uom_collection,  # Added UOM collection
)
from funtions import force_float_conversion

from .funtions import (
    build_filter,
    fetch_grn_data,
    fetch_inventory_map_for_day,
    fetch_opening_stock_for_day,
    fetch_purchase_items,
    fetch_store_dispatch_data,
    fetch_warehouse_return_data,
    get_dropdown_values,
    get_uom_precision_map,
    map_purchase_data,
    round_by_precision,
    safe_float,
)

from .models import ApproveItemRequest, ItemResponse, ListItemsResponse


router = APIRouter()

# ==================== UOM PRECISION HELPER ====================


# ==================== GET ENDPOINT - LIST ITEMS ====================
@router.get("/", response_model=ListItemsResponse)
async def get_all_items(request:Request,
    skip: int = Query(0, ge=0),
    limit: int = Query(10, le=50),
    varianceName: Optional[str] = None,
    itemName: Optional[str] = None,
    category: Optional[str] = None,
    subcategory: Optional[str] = None,
    locationName: str = Query(...),
    createdDate: Optional[str] = Query(
        None, description="Date (YYYY-MM-DD). Defaults to today."
    ),
    fetchDropdowns: bool = Query(True),
    categorySearch: Optional[str] = None,
    categoryPage: int = Query(1, ge=1),
    categoryLimit: int = Query(50, le=100),
    subcategorySearch: Optional[str] = None,
    subcategoryPage: int = Query(1, ge=1),
    subcategoryLimit: int = Query(50, le=100),
    varianceNameSearch: Optional[str] = None,
    varianceNamePage: int = Query(1, ge=1),
    varianceNameLimit: int = Query(50, le=100),
    itemNameSearch: Optional[str] = None,
    itemNamePage: int = Query(1, ge=1),
    itemNameLimit: int = Query(50, le=100),
    user = Depends(validate_token),
    permissions: dict = Depends(check_permission("yenerp","warehousephysicalstockvariancemodification","read"))
):
    tenant_id = request.state.tenant_id
    if not locationName:
        raise HTTPException(status_code=400, detail="locationName is required")

    if not createdDate:
        createdDate = ist_now().strftime("%Y-%m-%d")
    try:
        date_obj = datetime.strptime(createdDate, "%Y-%m-%d")
        next_day = date_obj + timedelta(days=1)
    except ValueError:
        raise HTTPException(
            status_code=400, detail="Invalid createdDate format. Use YYYY-MM-DD."
        )

    # Build full filter for main query
    full_filter = build_filter(varianceName, category, subcategory, itemName)

    # Fetch main items using full filter
    purchase_task = fetch_purchase_items(
        skip, limit, varianceName, category, subcategory, itemName
    )
    purchase_items, total_count, purchase_col = await purchase_task

    # Calculate previous day for opening stock BEFORE creating core_tasks
    prev_date_obj = date_obj - timedelta(days=1)

    core_tasks = [
        fetch_grn_data(date_obj, next_day),
        fetch_store_dispatch_data(date_obj),
        fetch_warehouse_return_data(date_obj, locationName),
        fetch_inventory_map_for_day(locationName, date_obj,tenant_id),
        fetch_opening_stock_for_day(locationName, date_obj,tenant_id),
    ]

    dropdown_tasks = []
    if fetchDropdowns:
        cat_base_filter = {
            k: v for k, v in full_filter.items() if k != "purchasecategoryName"
        }
        subcat_base_filter = {
            k: v for k, v in full_filter.items() if k != "purchasesubcategoryName"
        }
        variance_base_filter = {k: v for k, v in full_filter.items() if k != "itemName"}
        item_base_filter = {
            k: v for k, v in full_filter.items() if k != "itemgroupName"
        }

        dropdown_tasks = [
            get_dropdown_values(
                purchase_col,
                "purchasecategoryName",
                categoryPage,
                categoryLimit,
                categorySearch,
                base_filter=cat_base_filter,
            ),
            get_dropdown_values(
                purchase_col,
                "purchasesubcategoryName",
                subcategoryPage,
                subcategoryLimit,
                subcategorySearch,
                base_filter=subcat_base_filter,
            ),
            get_dropdown_values(
                purchase_col,
                "itemName",
                varianceNamePage,
                varianceNameLimit,
                varianceNameSearch,
                base_filter=variance_base_filter,
            ),
            get_dropdown_values(
                purchase_col,
                "itemgroupName",
                itemNamePage,
                itemNameLimit,
                itemNameSearch,
                base_filter=item_base_filter,
            ),
        ]

    all_tasks = core_tasks + dropdown_tasks
    results = await asyncio.gather(*all_tasks, return_exceptions=True)

    for i, r in enumerate(results):
        if isinstance(r, Exception):
            raise HTTPException(status_code=500, detail=f"Task {i} failed: {r}")

    # Unpack results - now opening_stock_map is part of core_tasks results
    (grn_data, store_map, warehouse_return_map, inventory_map, opening_stock_map) = (
        results[:5]
    )

    response = {
        "items": [],
        "totalItems": total_count,
        "currentDate": createdDate,
    }

    if (
        fetchDropdowns and len(results) >= 9
    ):  # Changed from 8 to 9 since we added opening_stock_map
        cat_dd, subcat_dd, variance_dd, item_dd = results[5:9]  # Changed indices
        response["dropdowns"] = {
            "category": {
                "items": cat_dd[0],
                "total": float(cat_dd[1]),
                "page": categoryPage,
                "limit": categoryLimit,
            },
            "subcategory": {
                "items": subcat_dd[0],
                "total": float(subcat_dd[1]),
                "page": subcategoryPage,
                "limit": subcategoryLimit,
            },
            "varianceName": {
                "items": variance_dd[0],
                "total": float(variance_dd[1]),
                "page": varianceNamePage,
                "limit": varianceNameLimit,
            },
            "itemName": {
                "items": item_dd[0],
                "total": float(item_dd[1]),
                "page": itemNamePage,
                "limit": itemNameLimit,
            },
        }

    formatted_items = await map_purchase_data(
        purchase_items,
        grn_data,
        store_map,
        opening_stock_map,  # Now correctly passed
        inventory_map,
        locationName,
        warehouse_return_map,
        date_obj,
    )

    msg = None
    if not grn_data and not store_map:
        msg = f"No GRN or store dispatch transactions found for {createdDate}."
    elif not grn_data:
        msg = f"No GRN transactions found for {createdDate}."
    elif not store_map:
        msg = f"No store dispatch transactions found for {createdDate}."

    response["items"] = [i.dict(by_alias=False) for i in formatted_items]
    response["totalItems"] = float(total_count)
    if msg:
        response["message"] = msg

    response = force_float_conversion(response)

    return response


def ist_today_range_utc():
    # IST midnight (00:00) in UTC
    ist_now = datetime.now(timezone.utc) + timedelta(hours=5, minutes=30)
    ist_midnight = ist_now.replace(hour=0, minute=0, second=0, microsecond=0)

    utc_start = ist_midnight - timedelta(hours=5, minutes=30)
    utc_end = utc_start + timedelta(days=1)

    return utc_start, utc_end


# ==================== PATCH ENDPOINT - APPROVE ITEM ====================
@router.patch("/{randomId}/approve")
async def approve_item(httprequest:Request,
    randomId: str,
    locationId: str = Query(..., description="Warehouse / Location ID"),
    status: str = Query("approved", description="Stock status - default: approved"),
    request: ApproveItemRequest = Body(...),
    user = Depends(validate_token),
    permissions: dict = Depends(check_permission("yenerp","warehousephysicalstockvariancemodification","approve"))
):
    tenant_id = request.state.tenant_id
    purchase_col = purchaseitem_collection()
    inventory_col = inventory_stock_collection(tenant_id)
    variance_col = approvedstocks_rm_collection(tenant_id)

    now = ist_now()

    random_id = randomId.strip()
    location_id = locationId.strip()

    # 1️⃣ Validate randomId exists in purchase collection
    purchase_doc = await purchase_col.find_one({"randomId": random_id})
    if not purchase_doc:
        raise HTTPException(
            status_code=404,
            detail=f"randomId {random_id} not found in Purchase collection",
        )

    # Get today IST range
    start_today, end_today = ist_today_range_utc()

    # 2️⃣ Fetch latest inventory record for today
    inventory_doc = await inventory_col.find_one(
        {
            "randomId": random_id,
            "locationId": location_id,
            "updatedAt": {
                "$gte": start_today,
                "$lt": end_today,
            },
        },
        sort=[("updatedAt", -1)],
    )

    if not inventory_doc:
        raise HTTPException(
            status_code=404,
            detail=f"No inventory record found for randomId={random_id} and locationId={location_id}",
        )

    current_status = inventory_doc.get("status", "pending")

    # 3️⃣ If already approved → stop
    if current_status.lower() == "approved":
        return {
            "message": "Already approved. Approval not required.",
            "randomId": random_id,
            "locationId": location_id,
            "status": current_status,
            "updatedAt": inventory_doc.get("updatedAt"),
        }

    # --- UOM Precision Logic ---
    uom_map = await get_uom_precision_map()
    uom_name = purchase_doc.get("uom", "")
    precision = uom_map.get(uom_name, 0)

    # 4️⃣ Get values ONLY from DB (no calculation)
    system_stock = round_by_precision(
        safe_float(inventory_doc.get("systemStock", 0)), precision
    )
    physical_stock = round_by_precision(
        safe_float(inventory_doc.get("physicalStock", 0)), precision
    )
    actual_variance = round_by_precision(
        safe_float(inventory_doc.get("variance", 0)), precision
    )

    # 5️⃣ If variance is zero → no approval needed
    if actual_variance == 0:
        return {
            "message": "No variance found. Approval not required.",
            "randomId": random_id,
            "locationId": location_id,
            "systemStock": system_stock,
            "physicalStock": physical_stock,
            "variance": 0,
            "status": "approved",
        }

    # 6️⃣ Update system stock using DB variance
    updated_system_stock = round_by_precision(system_stock + actual_variance, precision)
    updated_physical_stock = updated_system_stock  # make equal after approval

    # 7️⃣ Update inventory collection
    await inventory_col.update_one(
        {"_id": inventory_doc["_id"]},
        {
            "$set": {
                "previousSystemStock": system_stock,
                "systemStock": updated_system_stock,
                "physicalStock": physical_stock,
                "variance": 0,  # Reset after approval
                "status": status,
                "updatedAt": now,
                "lastUpdatedBy": request.approved_by,
            }
        },
    )

    # 8️⃣ Save variance history into approvedstocks collection
    variance_doc = {
        "randomId": random_id,
        "itemName": purchase_doc.get("itemName", ""),
        "locationId": location_id,
        "systemStockBefore": system_stock,
        "physicalClosing": physical_stock,
        "actualVariance": actual_variance,
        "systemStockAfter": updated_system_stock,
        "status": status,
        "approvedBy": request.approved_by,
        "description": request.description,
        "approvedAt": now,
    }

    insert_result = await variance_col.insert_one(variance_doc)

    variance_doc["_id"] = str(insert_result.inserted_id)

    # 9️⃣ Final Response
    return {
        "message": "Stock approved successfully",
        "data": variance_doc,
        "updatedSystemStock": updated_system_stock,
        "updatedPhysicalStock": updated_physical_stock,
        "varianceAfterApproval": 0,
        "status": status,
        "updatedAt": now.isoformat(),
    }


# ==================== GET APPROVED ITEMS ====================
@router.get("/approved")
async def get_approved_items(request:Request,
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1, le=100),
    locationId: Optional[str] = None,
    date: Optional[str] = None,
    status: str = Query("approved", description="Filter by approval status"),
    user = Depends(validate_token),
    permissions: dict = Depends(check_permission("yenerp","warehousephysicalstockvariancemodification","read"))
):
    tenant_id = request.state.tenant_id
    variance_col = approvedstocks_rm_collection(tenant_id)
    query = {"status": status}

    # Filter by locationId
    if locationId:
        query["locationId"] = locationId.strip()

    # Filter by approvedAt date
    if date:
        try:
            date_obj = datetime.strptime(date, "%Y-%m-%d")
            start = datetime(date_obj.year, date_obj.month, date_obj.day)
            end = start + timedelta(days=1)
            query["approvedAt"] = {"$gte": start, "$lt": end}
        except ValueError:
            raise HTTPException(
                status_code=400, detail="Invalid date format. Use YYYY-MM-DD"
            )

    skip = (page - 1) * limit

    total, items = await asyncio.gather(
        variance_col.count_documents(query),
        variance_col.find(query)
        
        .sort("approvedAt", -1)
        .skip(skip)
        .limit(limit)
        .to_list(limit),
    )

    if not items:
        return {
            "data": [],
            "page": page,
            "limit": limit,
            "total": total,
            "message": "No approved stock records found",
        }

    response = []

    for itm in items:
        response.append(
            {
                "id": str(itm["_id"]),
                "randomId": itm.get("randomId"),
                "itemName": itm.get("itemName"),
                "locationId": itm.get("locationId"),
                "systemStockBefore": itm.get("systemStockBefore"),
                "physicalClosing": itm.get("physicalClosing"),
                "actualVariance": itm.get("actualVariance"),
                "systemStockAfter": itm.get("systemStockAfter"),
                "status": itm.get("status", "approved"),
                "approvedBy": itm.get("approvedBy"),
                "description": itm.get("description"),
                "approvedAt": (
                    itm.get("approvedAt").strftime("%d-%m-%Y %H:%M:%S")
                    if itm.get("approvedAt")
                    else None
                ),
            }
        )

    return {
        "data": response,
        "page": page,
        "limit": limit,
        "total": total,
    }


@router.post("/dailystock/all")
async def create_daily_rm_snapshot_all_warehouses(request:Request,createdBy: str = "Inventory",user = Depends(validate_token),permissions: dict = Depends(check_permission("yenerp","warehousephysicalstockvariancemodification","add"))):
    tenant_id = request.state.tenant_id
    inventory_col = closingstocks_rm_collection(tenant_id)
    warehouse_col = warehouse_collection()
    purchase_items_col = purchaseitem_collection()
    stock_col = inventory_stock_collection(tenant_id)

    now = ist_now()
    today = now.strftime("%Y-%m-%d")

    warehouses = await warehouse_col.find({"status": 1, "type": "main"}).to_list(None)
    purchase_items = await purchase_items_col.find({"status": "active"}).to_list(None)

    if not warehouses or not purchase_items:
        raise HTTPException(
            status_code=404, detail="No active warehouses or items found"
        )

    # --- UOM Precision Map ---
    uom_map = await get_uom_precision_map()

    purchase_random_ids = [item["randomId"] for item in purchase_items]

    created_count = 0
    skipped_count = 0

    for wh in warehouses:
        warehouse_id = wh.get("warehouseId")
        if not warehouse_id:
            continue

        # Prevent duplicates for the day
        exists = await inventory_col.find_one(
            {"locationId": warehouse_id, "date": today}
        )

        if exists:
            skipped_count += 1
            continue

        stock_records = await stock_col.find(
            {"locationId": warehouse_id, "randomId": {"$in": purchase_random_ids}}
        ).to_list(None)

        stock_map = {rec["randomId"]: rec for rec in stock_records}

        # create per-item snapshot with all fields
        docs = []
        for item in purchase_items:
            rid = item["randomId"]
            stock = stock_map.get(rid, {})

            # Apply precision
            uom_name = item.get("uom", "")
            precision = uom_map.get(uom_name, 0)

            system_stock = round_by_precision(
                safe_float(stock.get("systemStock", 0.0)), precision
            )
            physical_stock = round_by_precision(
                safe_float(stock.get("physicalStock", 0.0)), precision
            )
            variance = round_by_precision(
                safe_float(stock.get("variance", 0.0)), precision
            )

            docs.append(
                {
                    "locationId": warehouse_id,
                    "randomId": rid,
                    "systemStock": system_stock,
                    "physicalStock": physical_stock,
                    "variance": variance,
                    "status": stock.get("status", "pending"),
                    "date": today,
                    "createdBy": createdBy,
                    "createdAt": now,
                }
            )

        await inventory_col.insert_many(docs)
        created_count += 1

    response = {
        "message": "Daily RM snapshot created for all warehouses",
        "created": created_count,
        "skipped": skipped_count,
    }
    return force_float_conversion(response)


@router.get("/items", response_model=List[ItemResponse])
async def get_item_names(
    page: int = Query(1, ge=1),
    limit: int = Query(10, le=50),
    search: str | None = None,
):
    collection = purchaseitem_collection()

    query = {}
    if search:
        query = {"itemName": {"$regex": search, "$options": "i"}}

    cursor = (
        collection.find(
            query,
            {
                "_id": 0,
                "randomId": 1,
                "itemName": 1,
            },
        )
        .skip((page - 1) * limit)
        .limit(limit)
    )

    items = []
    async for doc in cursor:
        items.append(
            {
                "randomId": doc["randomId"],
                "itemName": doc["itemName"],
            }
        )

    return items


@router.get("/stock-ledger")
async def stock_ledger_daily(request:Request,
    from_date: str = Query(...),
    to_date: str = Query(...),
    itemRandomId: Optional[str] = Query(None),
    locationName: Optional[str] = Query(None),
    user = Depends(validate_token),
    permissions: dict = Depends(check_permission("yenerp","warehousestockledger","read"))
):
    tenant_id = request.state.tenant_id
    try:
        start_date = datetime.strptime(from_date, "%Y-%m-%d")
        end_date = datetime.strptime(to_date, "%Y-%m-%d")

        if start_date > end_date:
            raise HTTPException(400, "Invalid date range")

        if not itemRandomId or not locationName:
            raise HTTPException(400, "itemRandomId and locationName are required")

        inventory_col = closingstocks_rm_collection(tenant_id)
        grn_col = grn_collection()
        store_col = storeDispatch_collection()
        item_col = purchaseitem_collection()  #  item master

        random_ids = [r.strip() for r in itemRandomId.split(",")]
        warehouse_id = locationName.strip()

        today = datetime.now().date()
        prev_day = start_date - timedelta(days=1)

        # ---------------------------------------------------
        # ITEM NAME MAP & UOM MAP
        # ---------------------------------------------------
        items = await item_col.find(
            {"randomId": {"$in": random_ids}},
            {"_id": 0, "randomId": 1, "itemName": 1, "uom": 1},
        ).to_list(None)

        item_map = {
            i["randomId"]: {"itemName": i.get("itemName", ""), "uom": i.get("uom", "")}
            for i in items
        }

        # Fetch UOM Precision
        uom_precision_map = await get_uom_precision_map()

        # ---------------------------------------------------
        # OPENING STOCK
        # ---------------------------------------------------
        opening_stocks = {rid: 0.0 for rid in random_ids}

        prev_docs = await inventory_col.find(
            {
                "date": prev_day.strftime("%Y-%m-%d"),
                "locationId": warehouse_id,
                "randomId": {"$in": random_ids},
            }
        ).to_list(None)

        for doc in prev_docs:
            rid = doc["randomId"]
            precision = uom_precision_map.get(item_map.get(rid, {}).get("uom", ""), 0)
            opening_stocks[rid] = round_by_precision(
                float(doc.get("systemStock", 0)), precision
            )

        # ---------------------------------------------------
        # LEDGER INIT
        # ---------------------------------------------------
        ledgers = {
            rid: {
                "randomId": rid,
                "itemDetails": item_map.get(rid, ""),
                "openingReference": {
                    "date": prev_day.strftime("%Y-%m-%d"),
                    "closingStock": opening_stocks[rid],
                },
                "transactions": [],
            }
            for rid in random_ids
        }

        balances = opening_stocks.copy()
        cur = start_date

        # ---------------------------------------------------
        # DATE LOOP
        # ---------------------------------------------------
        while cur <= end_date:
            next_day = cur + timedelta(days=1)

            snapshot_docs = await inventory_col.find(
                {
                    "date": cur.strftime("%Y-%m-%d"),
                    "locationId": warehouse_id,
                    "randomId": {"$in": random_ids},
                }
            ).to_list(None)

            snapshot_map = {}
            for d in snapshot_docs:
                rid = d["randomId"]
                precision = uom_precision_map.get(
                    item_map.get(rid, {}).get("uom", ""), 0
                )
                snapshot_map[rid] = round_by_precision(
                    float(d.get("systemStock", 0)), precision
                )

            # ---------- PAST DAYS ----------
            if cur.date() != today:
                for rid in random_ids:
                    closing_stock = snapshot_map.get(rid, balances[rid])

                    ledgers[rid]["transactions"].append(
                        {
                            "date": cur.strftime("%Y-%m-%d"),
                            "grnVendorName": "",
                            "dispatchBranch": "",
                            "returnedToVendor": "",
                            "inStock": 0,
                            "outStock": 0,
                            "returnedStock": 0,
                            "balanceStock": closing_stock,
                        }
                    )
                    balances[rid] = closing_stock

            # ---------- TODAY ----------
            else:
                grn_data = await grn_col.aggregate(
                    [
                        {
                            "$match": {
                                "grnDate": {"$gte": cur, "$lt": next_day},
                                "warehouseId": warehouse_id,
                            }
                        },
                        {"$unwind": "$itemDetails"},
                        {"$match": {"itemDetails.item_rand": {"$in": random_ids}}},
                        {
                            "$group": {
                                "_id": "$itemDetails.item_rand",
                                "received": {"$sum": "$itemDetails.receivedQuantity"},
                                "returned": {
                                    "$sum": {
                                        "$ifNull": ["$itemDetails.returnedQuantity", 0]
                                    }
                                },
                                "vendors": {"$addToSet": "$vendorName"},
                                "returnedByVendor": {
                                    "$push": {
                                        "vendor": "$vendorName",
                                        "returnedQty": {
                                            "$ifNull": [
                                                "$itemDetails.returnedQuantity",
                                                0,
                                            ]
                                        },
                                    }
                                },
                            }
                        },
                    ]
                ).to_list(None)

                grn_map = {g["_id"]: g for g in grn_data}

                dispatches = await store_col.find(
                    {
                        "type": "store",
                        "warehouseId": warehouse_id,
                        "date": {"$regex": f"^{cur.strftime('%Y-%m-%d')}"},
                    }
                ).to_list(None)

                out_map = {rid: 0.0 for rid in random_ids}
                branch_map = {rid: "" for rid in random_ids}

                for d in dispatches:
                    ids = d.get("randomId", [])
                    uoms = d.get("uom", [])
                    qtys = d.get("qty", [])
                    wts = d.get("weight", [])
                    branch = d.get("branchName", "")

                    if isinstance(ids, str):
                        ids = [ids]

                    for i, rid in enumerate(ids):
                        if rid in random_ids:
                            uom = uoms[i] if isinstance(uoms, list) else ""
                            qty = (
                                float(qtys[i])
                                if isinstance(qtys, list)
                                else float(qtys)
                            )
                            wt = float(wts[i]) if isinstance(wts, list) else float(wts)

                            out_map[rid] += (
                                wt
                                if uom.lower() in ["kg", "kgs", "ltr", "ltrs"]
                                else qty
                            )
                            branch_map[rid] = branch

                for rid in random_ids:
                    precision = uom_precision_map.get(
                        item_map.get(rid, {}).get("uom", ""), 0
                    )

                    received = round_by_precision(
                        float(grn_map.get(rid, {}).get("received", 0)), precision
                    )
                    returned = round_by_precision(
                        float(grn_map.get(rid, {}).get("returned", 0)), precision
                    )
                    out_stock = round_by_precision(out_map[rid], precision)

                    returned_vendors_info = ", ".join(
                        f"{r['vendor']}({r['returnedQty']})"
                        for r in grn_map.get(rid, {}).get("returnedByVendor", [])
                        if r.get("returnedQty", 0)
                    )

                    balances[rid] = balances[rid] + received - returned - out_stock
                    grn_entry = grn_map.get(rid, {})
                    if not isinstance(grn_entry, dict):
                        grn_entry = {}
                    uom = grn_entry.get("uom") or item_map.get(rid, {}).get("uom", "")

                    ledgers[rid]["transactions"].append(
                        {
                            "date": cur.strftime("%Y-%m-%d"),
                            "grnVendorName": ", ".join(
                                grn_map.get(rid, {}).get("vendors", [])
                            ),
                            "returnedToVendor": returned_vendors_info,
                            "dispatchBranch": branch_map.get(rid, ""),
                            "uom": uom,
                            "inStock": received,
                            "outStock": out_stock,
                            "returnedStock": returned,
                            "balanceStock": round_by_precision(
                                balances[rid], precision
                            ),
                        }
                    )

            cur = next_day

        # ---------------------------------------------------
        # CLOSING SUMMARY
        # ---------------------------------------------------
        for rid in random_ids:
            precision = uom_precision_map.get(item_map.get(rid, {}).get("uom", ""), 0)
            ledgers[rid]["closingSummary"] = {
                "date": end_date.strftime("%Y-%m-%d"),
                "closingStock": round_by_precision(balances[rid], precision),
            }

        return list(ledgers.values())

    except Exception as e:
        raise HTTPException(500, f"Ledger error: {str(e)}")


@router.get("/stock-ledger/excel")
async def stock_ledger_daily_excel(request:Request,
    from_date: str = Query(...),
    to_date: str = Query(...),
    itemRandomId: Optional[str] = Query(None),
    locationName: Optional[str] = Query(None),
    user = Depends(validate_token),
    permissions: dict = Depends(check_permission("yenerp","warehousestockledger","read"))
):
    tenant_id = request.state.tenant_id
    try:
        if not itemRandomId or not locationName:
            raise HTTPException(400, "itemRandomId and locationName are required")

        start_date = datetime.strptime(from_date, "%Y-%m-%d")
        end_date = datetime.strptime(to_date, "%Y-%m-%d")

        if start_date > end_date:
            raise HTTPException(400, "Invalid date range")

        inventory_col = closingstocks_rm_collection(tenant_id)
        grn_col = grn_collection()
        store_col = storeDispatch_collection()
        item_col = purchaseitem_collection()  # item master

        random_ids = [r.strip() for r in itemRandomId.split(",")]
        warehouse_name = locationName.strip()

        today = datetime.now().date()
        prev_day = start_date - timedelta(days=1)

        # -----------------------------
        # ITEM MAP & UOM MAP
        # -----------------------------
        items = await item_col.find(
            {"randomId": {"$in": random_ids}},
            {"_id": 0, "randomId": 1, "itemName": 1, "uom": 1},
        ).to_list(None)

        item_map = {
            i["randomId"]: {
                "itemName": i.get("itemName", i["randomId"]),
                "uom": i.get("uom", ""),
            }
            for i in items
        }

        # Fetch UOM Precision
        uom_precision_map = await get_uom_precision_map()

        # -----------------------------
        # OPENING STOCK
        # -----------------------------
        opening_stocks = {rid: 0.0 for rid in random_ids}
        prev_docs = await inventory_col.find(
            {
                "date": prev_day.strftime("%Y-%m-%d"),
                "locationId": warehouse_name,
                "randomId": {"$in": random_ids},
            }
        ).to_list(None)

        for doc in prev_docs:
            rid = doc["randomId"]
            precision = uom_precision_map.get(item_map.get(rid, {}).get("uom", ""), 0)
            opening_stocks[rid] = round_by_precision(
                float(doc.get("systemStock", 0)), precision
            )

        # -----------------------------
        # LEDGER INIT
        # -----------------------------
        ledgers = {
            rid: {
                "randomId": rid,
                "openingReference": {
                    "date": prev_day.strftime("%Y-%m-%d"),
                    "closingStock": opening_stocks[rid],
                },
                "transactions": [],
            }
            for rid in random_ids
        }

        balances = opening_stocks.copy()
        cur = start_date

        # -----------------------------
        # DATE LOOP
        # -----------------------------
        while cur <= end_date:
            next_day = cur + timedelta(days=1)

            # Get today's snapshot
            snapshot_docs = await inventory_col.find(
                {
                    "date": cur.strftime("%Y-%m-%d"),
                    "locationId": warehouse_name,
                    "randomId": {"$in": random_ids},
                }
            ).to_list(None)

            snapshot_map = {}
            for d in snapshot_docs:
                rid = d["randomId"]
                precision = uom_precision_map.get(
                    item_map.get(rid, {}).get("uom", ""), 0
                )
                snapshot_map[rid] = round_by_precision(
                    float(d.get("systemStock", 0)), precision
                )

            # PAST DAYS (no GRN/dispatch)
            if cur.date() != today:
                for rid in random_ids:
                    closing_stock = snapshot_map.get(rid, balances[rid])
                    ledgers[rid]["transactions"].append(
                        {
                            "date": cur.strftime("%Y-%m-%d"),
                            "grnVendorName": "",
                            "returnedToVendor": "",
                            "inStock": 0,
                            "outStock": 0,
                            "returnedStock": 0,
                            "dispatchBranch": "",
                            "balanceStock": closing_stock,
                        }
                    )
                    balances[rid] = closing_stock

            # TODAY
            else:
                # GRN aggregation
                grn_data = await grn_col.aggregate(
                    [
                        {
                            "$match": {
                                "grnDate": {"$gte": cur, "$lt": next_day},
                                "warehouseId": warehouse_name,
                            }
                        },
                        {"$unwind": "$itemDetails"},
                        {"$match": {"itemDetails.item_rand": {"$in": random_ids}}},
                        {
                            "$group": {
                                "_id": "$itemDetails.item_rand",
                                "received": {"$sum": "$itemDetails.receivedQuantity"},
                                "returned": {
                                    "$sum": {
                                        "$ifNull": ["$itemDetails.returnedQuantity", 0]
                                    }
                                },
                                "vendors": {"$addToSet": "$vendorName"},
                                "returnedByVendor": {
                                    "$push": {
                                        "vendor": "$vendorName",
                                        "returnedQty": {
                                            "$ifNull": [
                                                "$itemDetails.returnedQuantity",
                                                0,
                                            ]
                                        },
                                    }
                                },
                            }
                        },
                    ]
                ).to_list(None)

                grn_map = {g["_id"]: g for g in grn_data}

                # Dispatches
                dispatches = await store_col.find(
                    {
                        "type": "store",
                        "warehouseId": warehouse_name,
                        "date": {"$regex": f"^{cur.strftime('%Y-%m-%d')}"},
                    }
                ).to_list(None)

                out_map = {rid: 0.0 for rid in random_ids}
                branch_map = {rid: "" for rid in random_ids}

                for d in dispatches:
                    ids = d.get("randomId", [])
                    uoms = d.get("uom", [])
                    qtys = d.get("qty", [])
                    wts = d.get("weight", [])
                    branch = d.get("branchName", "")

                    if isinstance(ids, str):
                        ids = [ids]

                    for i, rid in enumerate(ids):
                        if rid in random_ids:
                            uom = uoms[i] if isinstance(uoms, list) else ""
                            qty = (
                                float(qtys[i])
                                if isinstance(qtys, list)
                                else float(qtys)
                            )
                            wt = float(wts[i]) if isinstance(wts, list) else float(wts)

                            out_map[rid] += (
                                wt
                                if uom.lower() in ["kg", "kgs", "ltr", "ltrs"]
                                else qty
                            )
                            branch_map[rid] = branch

                # Ledger update
                for rid in random_ids:
                    precision = uom_precision_map.get(
                        item_map.get(rid, {}).get("uom", ""), 0
                    )

                    received = round_by_precision(
                        float(grn_map.get(rid, {}).get("received", 0)), precision
                    )
                    returned = round_by_precision(
                        float(grn_map.get(rid, {}).get("returned", 0)), precision
                    )
                    out_stock = round_by_precision(out_map[rid], precision)

                    balances[rid] = balances[rid] + received - returned - out_stock

                    returned_vendors_info = ", ".join(
                        f"{r['vendor']}({r['returnedQty']})"
                        for r in grn_map.get(rid, {}).get("returnedByVendor", [])
                        if r.get("returnedQty", 0)
                    )

                    ledgers[rid]["transactions"].append(
                        {
                            "date": cur.strftime("%Y-%m-%d"),
                            "grnVendorName": ", ".join(
                                grn_map.get(rid, {}).get("vendors", [])
                            ),
                            "returnedToVendor": returned_vendors_info,
                            "inStock": received,
                            "outStock": out_stock,
                            "returnedStock": returned,
                            "dispatchBranch": branch_map.get(rid, ""),
                        }
                    )

            cur = next_day

        # -----------------------------
        # CLOSING SUMMARY
        # -----------------------------
        for rid in random_ids:
            precision = uom_precision_map.get(item_map.get(rid, {}).get("uom", ""), 0)
            ledgers[rid]["closingSummary"] = {
                "date": end_date.strftime("%Y-%m-%d"),
                "closingStock": round_by_precision(balances[rid], precision),
            }

        # -----------------------------
        # Excel Export
        # -----------------------------
        wb = Workbook()
        wb.remove(wb.active)
        header_fill = PatternFill("solid", fgColor="4F81BD")
        header_font = Font(color="FFFFFF", bold=True)

        for rid, data in ledgers.items():
            item_name = item_map.get(rid, {}).get("itemName", rid)
            uom = item_map.get(rid, {}).get("uom", "")

            ws = wb.create_sheet(title=item_name[:31])

            ws.append([f"Item Name: {item_name}"])
            ws.append([f"Location: {locationName}"])
            ws.append([f"Date Range: {from_date} to {to_date}"])
            ws.append([f"UOM: {uom}"])
            ws.append([])

            ws.append(
                [
                    "Date",
                    "GRN Vendor",
                    "In Stock",
                    "Returned to Vendor",
                    "Returned Stock",
                    "Dispatch Branch",
                    "Out Stock",
                    "Balance Stock",
                    "UOM",
                ]
            )
            for col in range(1, 10):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.fill = header_fill
                cell.font = header_font

            for t in data["transactions"]:
                ws.append(
                    [
                        t.get("date", ""),
                        t.get("grnVendorName", ""),
                        t.get("inStock", 0),
                        t.get("returnedToVendor", ""),
                        t.get("returnedStock", 0),
                        t.get("dispatchBranch", ""),
                        t.get("outStock", 0),
                        t.get("balanceStock", 0),
                        uom,
                    ]
                )

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"StockLedger_Daily_{from_date}_to_{to_date}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except Exception as e:
        raise HTTPException(500, f"Excel export error: {str(e)}")
