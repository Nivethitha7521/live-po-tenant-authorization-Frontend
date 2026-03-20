import io
from openpyxl import Workbook
from datetime import datetime, timedelta
from typing import List, Optional
from fastapi import APIRouter, HTTPException, Query, Response
import pandas as pd
from fastapi import Depends
from dependencies.auth import validate_token
from middlewares.permission_middleware import check_permission
from ApInvoiceReport.models import DropdownResponse
from .models import PaginatedResponse, PaginatedSales, Sales
from db.collections import dayEnd, location
from .utils import transform_day_end
from fastapi.responses import StreamingResponse
from io import BytesIO
from openpyxl.utils import get_column_letter

router = APIRouter()

collection = dayEnd
location_collection = location


@router.get("/date-dropdown", response_model=DropdownResponse)
async def get_apinvoice_endpoint( user=Depends(validate_token),
    permissions=Depends(check_permission("yenerp", "posreport", "read"))):
    # --- 1. Extract year, month, day from dayClosingDateTime ---
    pipeline_dates = [
        {"$match": {"dayClosingDateTime": {"$exists": True}}},
        {
            "$group": {
                "_id": None,
                "years": {"$addToSet": {"$year": "$dayClosingDateTime"}},
                "months": {"$addToSet": {"$month": "$dayClosingDateTime"}},
                "days": {"$addToSet": {"$dayOfMonth": "$dayClosingDateTime"}},
            }
        },
    ]
    date_result = await collection.aggregate(pipeline_dates).to_list(1)

    years = sorted(map(str, date_result[0]["years"])) if date_result else []
    months = sorted(f"{m:02d}" for m in date_result[0]["months"]) if date_result else []
    days = sorted(date_result[0]["days"]) if date_result else []

    return DropdownResponse(
        yearIn=years,
        monthIn=months,
        daysIn=days,
    )


@router.get("/report", response_model=PaginatedResponse)
async def get_dayEnd_reports(
    branchName: Optional[List[str]] = Query(None),  # FIX: Changed to List[str] for multiple
    startDate: Optional[str] = Query(None, description="YYYY-MM-DD"),
    endDate: Optional[str] = Query(None, description="YYYY-MM-DD"),
    page: int = 1,
    limit: int = 10,
     user=Depends(validate_token),
    permissions=Depends(check_permission("yenerp", "posreport", "read"))
):
    filter_query = {}

    # FIX: Use $in operator for multiple branch names
    if branchName:
        filter_query["locationId"] = {"$in": branchName}

    # Date filter
    if startDate:
        try:
            start = datetime.strptime(startDate, "%Y-%m-%d")
            # If endDate is provided, use it; otherwise range is just the start date
            end = (
                datetime.strptime(endDate, "%Y-%m-%d") + timedelta(days=1)
                if endDate
                else start + timedelta(days=1)
            )
            filter_query["dayClosingDateTime"] = {"$gte": start, "$lt": end}
        except ValueError:
            raise HTTPException(400, "Invalid date format. Use YYYY-MM-DD.")

    cursor = collection.find(filter_query).sort("dayClosingDateTime", -1)
    raw_docs = await cursor.to_list(length=None)

    # Transform documents → rows
    all_rows = []
    for doc in raw_docs:
        all_rows.extend(transform_day_end(doc))

    # Pagination AFTER transformation
    totalrecords = len(all_rows)
    skip = (page - 1) * limit
    paginated_rows = all_rows[skip : skip + limit]
    totalpages = (totalrecords + limit - 1) // limit

    return {
        "page": page,
        "limit": limit,
        "skip": skip,
        "totalrecords": totalrecords,
        "totalpages": totalpages,
        "items": paginated_rows,
    }


@router.get("/export")
async def download_dayEnd_reports_excel(
    branchName: Optional[List[str]] = Query(None),  # FIX: Changed to List[str] for multiple
    startDate: Optional[str] = Query(None, description="YYYY-MM-DD"),
    endDate: Optional[str] = Query(None, description="YYYY-MM-DD"),
     user=Depends(validate_token),
    permissions=Depends(check_permission("yenerp", "posreport", "read"))
):
    filter_query = {}

    # FIX: Use $in operator for multiple branch names
    if branchName:
        filter_query["locationId"] = {"$in": branchName}

    # Date filter
    if startDate:
        try:
            start = datetime.strptime(startDate, "%Y-%m-%d")
            end = (
                datetime.strptime(endDate, "%Y-%m-%d") + timedelta(days=1)
                if endDate
                else start + timedelta(days=1)
            )
            filter_query["dayClosingDateTime"] = {"$gte": start, "$lt": end}
        except ValueError:
            raise HTTPException(400, "Invalid date format. Use YYYY-MM-DD.")

    # Fetch Data
    cursor = collection.find(filter_query).sort("dayClosingDateTime", -1)
    raw_docs = await cursor.to_list(length=None)

    # Transform for Excel
    all_rows = []
    for doc in raw_docs:
        all_rows.extend(transform_day_end(doc))

    if not all_rows:
        raise HTTPException(status_code=404, detail="No records found")

    # Create DataFrame
    df = pd.DataFrame(all_rows)
    
    # Ensure columns exist
    expected_cols = [
        "Date", "Time", "Branch", "Type", "Cash", "Card", "UPI", 
        "Others/Online", "Total Amount"
    ]
    # Reindex to ensure order, filling missing with empty string
    df = df.reindex(columns=expected_cols, fill_value="")

    # Create Excel in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")

    output.seek(0)
    download_time = datetime.now().strftime("%d-%m-%Y_%H-%M")
    filename = f"DayEnd_YenERP_{download_time}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/overallsales/report", response_model=PaginatedSales)
async def get_sales_reports(
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1, le=200),
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    branchName: Optional[List[str]] = Query(None), # Already correct
     user=Depends(validate_token),
    permissions=Depends(check_permission("yenerp", "posreport", "read"))
):
    filters = {}

    # ------------ Date Filter ------------
    date_filter = {}
    if start_date:
        date_filter["$gte"] = datetime.combine(start_date.date(), datetime.min.time())
    if end_date:
        date_filter["$lte"] = datetime.combine(end_date.date(), datetime.max.time())
    if date_filter:
        filters["dayClosingDateTime"] = date_filter

    # ------------ Branch Filter ------------
    if branchName:
        filters["locationId"] = {"$in": branchName}

    # ------------ Total Count ------------
    totalcount = await dayEnd.count_documents(filters)

    # ------------ Pagination ------------
    skip = (page - 1) * limit
    cursor = dayEnd.find(filters).skip(skip).limit(limit)

    # ------------ Get branch randomId ------------
    branch_docs = await location_collection.find({}, {"branchName": 1, "randomId": 1}).to_list(None)
    branch_lookup = {b["branchName"]: b.get("randomId") for b in branch_docs}

    items = []
    async for doc in cursor:
        doc["_id"] = str(doc["_id"])
        doc["randomId"] = branch_lookup.get(doc.get("branchName"), None)

        if "dayClosingDateTime" in doc and isinstance(doc["dayClosingDateTime"], datetime):
            doc["dayClosingDateTime"] = doc["dayClosingDateTime"].strftime("%Y-%m-%d")

        items.append(Sales(**doc))

    totalpages = (totalcount + limit - 1) // limit

    return PaginatedSales(
        page=page,
        limit=limit,
        totalrecords=totalcount,
        totalpages=totalpages,
        items=items,
    )


@router.get("/overallsales/export")
async def export_sales_excel(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    branchName: Optional[List[str]] = Query(None), # Already correct
     user=Depends(validate_token),
    permissions=Depends(check_permission("yenerp", "posreport", "read"))
):
    filters = {}

    # --------------------------
    # Date Filter
    # --------------------------
    if start_date or end_date:
        date_filter = {}
        if start_date:
            date_filter["$gte"] = datetime.combine(start_date.date(), datetime.min.time())
        if end_date:
            date_filter["$lte"] = datetime.combine(end_date.date(), datetime.max.time())
        filters["dayClosingDateTime"] = date_filter

    # --------------------------
    # Branch Filter
    # --------------------------
    if branchName:
        filters["locationId"] = {"$in": branchName}

    # --------------------------
    # Fetch data from MongoDB
    # --------------------------
    docs = await collection.find(filters).to_list(None)

    if not docs:
         raise HTTPException(status_code=404, detail="No records found")

    # Fetch branch → randomId mapping
    branch_docs = await location_collection.find({}, {"branchName": 1, "randomId": 1}).to_list(None)
    branch_lookup = {b["branchName"]: b.get("randomId") for b in branch_docs}

    # --------------------------
    # Excel Workbook
    # --------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = [
        "Code", "Location", "Card", "Cash", "UPI", "Others", "Sales", 
        "Dinning", "SalesInvoice", "SalesOrders", "SaleDate",
    ]
    ws.append(headers)

    # --------------------------
    # Write rows
    # --------------------------
    for d in docs:
        branch = d.get("branchName")
        random_id = branch_lookup.get(branch)

        day_date = (
            d.get("dayClosingDateTime").strftime("%Y-%m-%d")
            if d.get("dayClosingDateTime")
            else ""
        )

        row = [
            random_id,
            branch,
            (d.get("systemCardSales")),
            (d.get("systemCashSales")),
            (d.get("systemUpiSales")),
            (d.get("systemOtherSales")),
            (d.get("totalSystemSales")),
            (d.get("totalKotSales")),
            (d.get("totalTakeAwaySales")),
            (d.get("totalSaleOrderSales")),
            day_date,
        ]
        ws.append(row)

    # --------------------------
    # Auto column width
    # --------------------------
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_len + 2

    # --------------------------
    # Return Excel file
    # --------------------------
    download_time = datetime.now().strftime("%d-%m-%Y_%H-%M")
    filename = f"Sales_YenERP_{download_time}.xlsx"
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )