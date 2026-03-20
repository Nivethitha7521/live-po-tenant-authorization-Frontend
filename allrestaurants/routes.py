from datetime import datetime, timedelta
import io
import re
import tempfile
from fastapi import APIRouter, HTTPException, Query
from typing import List, Optional
from fastapi import Depends
from dependencies.auth import validate_token
from middlewares.permission_middleware import check_permission
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font

from .models import (
    BranchData,
    AllRestaurant,
    DateDropdownResponse,
)
from db.collections import invoices

router = APIRouter()

NUMERIC_FIELDS = [
    "myAmount",
    "discountAmount",
    "netAmount",
    "deliveryCharge",
    "containerCharge",
    "serviceCharge",
    "additionalCharge",
    "totalTax",
    "roundOff",
    "waivedoff",
    "totalAmount",
    "onlineTaxCalculated",
    "gstPaidByMerchant",
    "gstPaidByEcommerce",
    "cash",
    "card",
    "upi",
    "duePayment",
    "others",
    "wallet",
    "online",
    "pax",
]


@router.get("/date-dropdown", response_model=DateDropdownResponse)
async def get_dispatch_date_dropdown( user=Depends(validate_token),
    permissions=Depends(check_permission("yenerp", "posreport", "read"))):
    collection = invoices

    pipeline = [
        {"$match": {"invoiceDateTime": {"$type": "date"}}},
        {
            "$group": {
                "_id": None,
                "years": {"$addToSet": {"$year": "$invoiceDateTime"}},
                "months": {"$addToSet": {"$month": "$invoiceDateTime"}},
                "days": {"$addToSet": {"$dayOfMonth": "$invoiceDateTime"}},
            }
        },
    ]

    result = await collection.aggregate(pipeline).to_list(1)

    if not result:
        return DateDropdownResponse(yearIn=[], monthIn=[], daysIn=[])

    return DateDropdownResponse(
        yearIn=sorted(map(str, result[0]["years"])),
        monthIn=sorted(f"{m:02d}" for m in result[0]["months"]),
        daysIn=sorted(result[0]["days"]),
    )


@router.get("/report")
async def get_restaurants(
    startDate: Optional[datetime] = Query(None),
    endDate: Optional[datetime] = Query(None),
    branchName: Optional[List[str]] = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1),
     user=Depends(validate_token),
    permissions=Depends(check_permission("yenerp", "posreport", "read"))
):
    collection = invoices

    # -----------------------------
    # Filters
    # -----------------------------
    match_filter = {}
    if startDate:
        startDate = datetime.combine(startDate.date(), datetime.min.time())
        match_filter.setdefault("invoiceDateTime", {})["$gte"] = startDate
    if endDate:
        endDate = datetime.combine(endDate.date(), datetime.min.time()) + timedelta(
            days=1
        )
        match_filter.setdefault("invoiceDateTime", {})["$lt"] = endDate

    if branchName:
        match_filter["locationId"] = {"$in": branchName}

    # -----------------------------
    # Aggregation with Facet for Pagination
    # -----------------------------
    pipeline = [
        {"$match": match_filter},
        {
            "$facet": {
                # 1. Metadata: Total count of unique branches
                "metadata": [
                    {"$group": {"_id": "$branchName"}},
                    {"$count": "total_records"},
                ],
                # 2. Data: Branch-level aggregation with pagination
                "data": [
                    {
                        "$group": {
                            "_id": "$branchName",
                            "first_invoice_no": {"$min": "$invoiceNo"},
                            "last_invoice_no": {"$max": "$invoiceNo"},
                            "last_invoice_date": {"$max": "$invoiceDateTime"},
                            "total_no_of_bills": {"$sum": 1},
                            **{
                                field: {"$sum": {"$ifNull": [f"${field}", 0]}}
                                for field in NUMERIC_FIELDS
                            },
                        }
                    },
                    {"$sort": {"_id": 1}},
                    {"$skip": (page - 1) * limit},
                    {"$limit": limit},
                    {
                        "$project": {
                            "_id": 0,
                            "branchName": "$_id",
                            "invoiceNo": {
                                "$concat": [
                                    {"$toString": "$first_invoice_no"},
                                    " - ",
                                    {"$toString": "$last_invoice_no"},
                                ]
                            },
                            "total_no_of_bills": 1,
                            **{
                                field: {"$round": [f"${field}", 2]}
                                for field in NUMERIC_FIELDS
                            },
                            "invoiceDate": {
                                "$dateToString": {
                                    "format": "%Y-%m-%d %H:%M:%S",
                                    "date": "$last_invoice_date",
                                }
                            },
                        }
                    },
                ],
            }
        },
    ]

    result = await collection.aggregate(pipeline).to_list(1)
    result = result[0] if result else {"metadata": [], "data": []}

    # Extract metadata
    meta = result["metadata"][0] if result["metadata"] else {"total_records": 0}
    total_records = meta["total_records"]
    total_pages = (total_records + limit - 1) // limit if limit > 0 else 0

    items = [AllRestaurant(**row) for row in result["data"]]

    # -----------------------------
    # Return response (Items only, no summary)
    # -----------------------------
    return {
        "totalcount": total_records,
        "totalpages": total_pages,
        "page": page,
        "limit": limit,
        "items": items,
    }


@router.get("/export")
async def export_restaurants_excel(
    startDate: Optional[datetime] = Query(None),
    endDate: Optional[datetime] = Query(None),
    branchName: Optional[List[str]] = Query(None),
     user=Depends(validate_token),
    permissions=Depends(check_permission("yenerp", "posreport", "read"))
):
    """
    Export complete restaurant report with ALL fields to Excel, including Summary.
    """
    collection = invoices

    # -----------------------------
    # Filters
    # -----------------------------
    match_filter = {}
    start_date_filter = startDate
    end_date_filter = endDate

    if startDate:
        startDate = datetime.combine(startDate.date(), datetime.min.time())
        match_filter.setdefault("invoiceDateTime", {})["$gte"] = startDate
    if endDate:
        endDate = datetime.combine(endDate.date(), datetime.min.time()) + timedelta(
            days=1
        )
        match_filter.setdefault("invoiceDateTime", {})["$lt"] = endDate

    if branchName:
        match_filter["locationId"] = {"$in": branchName}

    # -----------------------------
    # 1. Fetch All Branch Data (No Pagination)
    # -----------------------------
    detailed_pipeline = [
        {"$match": match_filter},
        {
            "$group": {
                "_id": "$branchName",
                "first_invoice_no": {"$min": "$invoiceNo"},
                "last_invoice_no": {"$max": "$invoiceNo"},
                "last_invoice_date": {"$max": "$invoiceDateTime"},
                "total_no_of_bills": {"$sum": 1},
                **{
                    field: {"$sum": {"$ifNull": [f"${field}", 0]}}
                    for field in NUMERIC_FIELDS
                },
            }
        },
        {"$sort": {"_id": 1}},
        {
            "$project": {
                "_id": 0,
                "branchName": "$_id",
                "invoiceNo": {
                    "$concat": [
                        {"$toString": "$first_invoice_no"},
                        " - ",
                        {"$toString": "$last_invoice_no"},
                    ]
                },
                "total_no_of_bills": 1,
                **{field: {"$round": [f"${field}", 2]} for field in NUMERIC_FIELDS},
                "invoiceDate": {
                    "$dateToString": {
                        "format": "%Y-%m-%d %H:%M:%S",
                        "date": "$last_invoice_date",
                    }
                },
            }
        },
    ]

    rows_data = await collection.aggregate(detailed_pipeline).to_list(length=None)

    # Convert to Models for consistent attribute access in export function
    branches = [
        BranchData(branchName=row["branchName"], rows=[AllRestaurant(**row)])
        for row in rows_data
    ]

    # -----------------------------
    # 2. Calculate Summary
    # -----------------------------
    branch_totals_pipeline = [
        {"$match": match_filter},
        {
            "$group": {
                "_id": "$branchName",
                "total_no_of_bills": {"$sum": 1},
                **{
                    field: {"$sum": {"$ifNull": [f"${field}", 0]}}
                    for field in NUMERIC_FIELDS
                },
            }
        },
    ]
    branch_totals = await collection.aggregate(branch_totals_pipeline).to_list(
        length=None
    )

    summary = {"TOTAL": {}, "MIN": {}, "MAX": {}, "AVG": {}}

    # Initialize totals
    for field in NUMERIC_FIELDS:
        summary["TOTAL"][field] = 0

    for branch in branch_totals:
        summary["TOTAL"]["total_no_of_bills"] = (
            summary["TOTAL"].get("total_no_of_bills", 0) + branch["total_no_of_bills"]
        )
        for field in NUMERIC_FIELDS:
            summary["TOTAL"][field] += branch.get(field, 0)

    for field in NUMERIC_FIELDS + ["total_no_of_bills"]:
        values = [branch.get(field, 0) for branch in branch_totals]
        summary["MIN"][field] = round(min(values), 2) if values else 0
        summary["MAX"][field] = round(max(values), 2) if values else 0
        summary["AVG"][field] = round(sum(values) / len(values), 2) if values else 0

    # Create a temporary report object to pass to the helper
    class TempReport:
        def __init__(self, summary, branches):
            self.summary = summary
            self.branches = branches

    report = TempReport(summary, branches)

    # -----------------------------
    # 3. Generate Excel
    # -----------------------------
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        filename = tmp.name

    export_report_complete(report, filename, start_date_filter, end_date_filter)

    return FileResponse(
        path=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"AllRestaurant_YenERP.xlsx",
    )


def export_report_complete(report, filename, startDate=None, endDate=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Date range header
    date_range = ""
    if startDate and endDate:
        date_range = (
            f"{startDate.strftime('%Y-%m-%d')} to {endDate.strftime('%Y-%m-%d')}"
        )
    elif startDate:
        date_range = f"From {startDate.strftime('%Y-%m-%d')}"
    elif endDate:
        date_range = f"Up to {endDate.strftime('%Y-%m-%d')}"
    ws.append(["Date:", date_range])
    ws.append(["Name:", "All Restaurant Sales Report"])
    ws.append([])

    # Column headers
    columns = [
        "Restaurants",
        "Invoice Nos.",
        "Total no. of bills",
        "My Amount",
        "Total Discount",
        "Net Sales(M.A - T.D)",
        "Delivery Charge",
        "Container Charge",
        "Service Charge",
        "Additional Charge",
        "Total Tax",
        "Round Off",
        "Waived off",
        "Total Sales",
        "Online Tax Calculated",
        "GST Paid by Merchant",
        "GST Paid by Ecommerce",
        "Cash",
        "Card",
        "Due Payment",
        "Other",
        "Wallet",
        "Online",
        "Pax",
        "Data Synced",
    ]

    field_mapping = {
        "Total no. of bills": "total_no_of_bills",
        "My Amount": "totalAmount",
        "Total Discount": "discountAmount",
        "Net Sales(M.A - T.D)": "netAmount",
        "Delivery Charge": "deliveryCharge",
        "Container Charge": "containerCharge",
        "Service Charge": "serviceCharge",
        "Additional Charge": "additionalCharge",
        "Total Tax": "totalTax",
        "Round Off": "roundOff",
        "Waived off": "waivedoff",
        "Total Sales": "totalAmount",
        "Online Tax Calculated": "onlineTaxCalculated",
        "GST Paid by Merchant": "gstPaidByMerchant",
        "GST Paid by Ecommerce": "gstPaidByEcommerce",
        "Cash": "cash",
        "Card": "card",
        "Due Payment": "duePayment",
        "Other": "others",
        "Wallet": "wallet",
        "Online": "online",
        "Pax": "pax",
    }

    # Append header row
    ws.append(columns)
    header_row = ws.max_row

    # Make header bold and black
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="000000")

    # --- Write summary rows (TOTAL, MIN, MAX, AVG) ---
    summary_keys = ["TOTAL", "MIN", "MAX", "AVG"]
    display_keys = ["Total", "Min.", "Max.", "Avg."]

    for i, key in enumerate(summary_keys):
        summary_dict = report.summary.get(key, {})
        row = [display_keys[i], ""]
        for col in columns[2:-1]:
            field_name = field_mapping.get(col)
            value = summary_dict.get(field_name, 0) if field_name else 0
            row.append(round(value, 2) if isinstance(value, (int, float)) else value)
        row.append("")
        ws.append(row)

    # --- Write branch rows ---
    for branch in report.branches:
        for row_data in branch.rows:
            data_row = [branch.branchName, getattr(row_data, "invoiceNo", "")]
            for col in columns[2:-1]:
                field_name = field_mapping.get(col)
                value = getattr(row_data, field_name, 0) if field_name else 0
                data_row.append(
                    round(value, 2) if isinstance(value, (int, float)) else value
                )

            # Data Synced
            data_synced = getattr(row_data, "invoiceDate", "")
            if hasattr(row_data, "invoiceDate") and row_data.invoiceDate:
                if isinstance(row_data.invoiceDate, datetime):
                    data_synced = row_data.invoiceDate.strftime("%Y-%m-%d %H:%M:%S")
                else:
                    data_synced = str(row_data.invoiceDate)
            data_row.append(data_synced)
            ws.append(data_row)

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = list(col)
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        col_letter = column[0].column_letter
        ws.column_dimensions[col_letter].width = max_length + 2

    # Freeze header row
    ws.freeze_panes = f"A{header_row + 1}"

    wb.save(filename)


@router.get("/Paymode/report", response_model=dict)
async def get_itemwisesales_by_date_range(
    startDate: Optional[datetime] = Query(None),
    endDate: Optional[datetime] = Query(None),
    branchName: Optional[List[str]] = Query(None),
    employeeName: Optional[List[str]] = Query(None),
    customerNumber: Optional[List[str]] = Query(None),
    invoiceId: Optional[List[str]] = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1), user=Depends(validate_token),
    permissions=Depends(check_permission("yenerp", "posreport", "read"))
):
    try:
        collection = invoices

        # ---------------- FILTER ----------------
        query = {"salesType": "Dinning", "status": "active"}

        if startDate and endDate:
            query["invoiceDateTime"] = {
                "$gte": startDate,
                "$lte": endDate.replace(hour=23, minute=59, second=59),
            }
        elif startDate:
            query["invoiceDateTime"] = {"$gte": startDate}
        elif endDate:
            query["invoiceDateTime"] = {
                "$lte": endDate.replace(hour=23, minute=59, second=59)
            }

        if branchName:
            query["locationId"] = {"$in": branchName}

        if employeeName:
            query["employeeName"] = {"$in": employeeName}

        if customerNumber:
            query["customerPhoneNumber"] = {"$in": customerNumber}

        if invoiceId:
            query["invoiceId"] = {"$in": invoiceId}

        skip = (page - 1) * limit

        # ---------------- PIPELINE ----------------
        pipeline = [
            {"$match": query},
            {"$unwind": {"path": "$varianceName", "includeArrayIndex": "idx"}},
            {
                "$project": {
                    "_id": 1,
                    "itemName": {"$arrayElemAt": ["$itemName", "$idx"]},
                    "varianceName": 1,
                    "price": {"$arrayElemAt": ["$price", "$idx"]},
                    "qty": {"$arrayElemAt": ["$qty", "$idx"]},
                    "weight": {"$arrayElemAt": ["$weight", "$idx"]},
                    "amount": {"$arrayElemAt": ["$amount", "$idx"]},
                    "tax": {"$arrayElemAt": ["$tax", "$idx"]},
                    "uom": {"$arrayElemAt": ["$uom", "$idx"]},
                    "salesType": 1,
                    "invoiceNo": 1,
                    "customerNumber": 1,
                    "totalAmount": 1,
                    "branchName": 1,
                    "invoiceId": 1,
                    "employeeName": 1,
                    "paymentType": 1,
                    "invoiceDateTime": 1,
                    "status": 1,
                    # ✅ FIXED netAmount
                    "netAmount": {
                        "$round": [
                            {
                                "$multiply": [
                                    {
                                        "$ifNull": [
                                            {"$arrayElemAt": ["$price", "$idx"]},
                                            0,
                                        ]
                                    },
                                    {
                                        "$ifNull": [
                                            {"$arrayElemAt": ["$qty", "$idx"]},
                                            0,
                                        ]
                                    },
                                ]
                            },
                            2,
                        ]
                    },
                }
            },
            {"$sort": {"invoiceDateTime": -1}},  # ✅ FIXED sorting
            {"$skip": skip},
            {"$limit": limit},
        ]

        cursor = collection.aggregate(pipeline)

        items = []
        async for doc in cursor:
            doc["_id"] = str(doc["_id"])
            items.append(doc)

        # ---------------- COUNT (FIXED) ----------------
        count_pipeline = [
            {"$match": query},
            {"$unwind": "$varianceName"},
            {"$count": "total"},
        ]

        count_result = await collection.aggregate(count_pipeline).to_list(1)
        total_docs = count_result[0]["total"] if count_result else 0

        return {
            "totalcount": total_docs,
            "totalpages": (total_docs + limit - 1) // limit,
            "page": page,
            "limit": limit,
            "items": items,
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/Paymode/export")
async def get_paymode_excel(
    startDate: Optional[datetime] = Query(None),
    endDate: Optional[datetime] = Query(None),
    branchName: Optional[List[str]] = Query(None),
    employeeName: Optional[List[str]] = Query(None),
    customerNumber: Optional[List[str]] = Query(None),
    invoiceId: Optional[List[str]] = Query(None), user=Depends(validate_token),
    permissions=Depends(check_permission("yenerp", "posreport", "read"))
):
    try:
        collection = invoices

        # ---------------- FILTER ----------------
        query = {"salesType": "Dinning", "status": "active"}

        if startDate and endDate:
            query["invoiceDateTime"] = {
                "$gte": startDate,
                "$lte": endDate.replace(hour=23, minute=59, second=59),
            }
        elif startDate:
            query["invoiceDateTime"] = {"$gte": startDate}
        elif endDate:
            query["invoiceDateTime"] = {
                "$lte": endDate.replace(hour=23, minute=59, second=59)
            }

        if branchName:
            query["locationId"] = {"$in": branchName}

        if employeeName:
            query["employeeName"] = {"$in": employeeName}

        if customerNumber:
            query["customerPhoneNumber"] = {"$in": customerNumber}

        if invoiceId:
            query["invoiceId"] = {"$in": invoiceId}

        cursor = collection.find(query).sort("invoiceDateTime", -1)

        # ---------------- EXCEL ----------------
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        headers = [
            "Branch Name",
            "Invoice No",
            "GST No",
            "Invoice Date",
            "KOT No",
            "Payment Type",
            "Payment Description",
            "Sales Type",
            "Status",
            "Sub Order Type",
            "Area",
            "Virtual Branch",
            "assign_to",
            "Group Name",
            "Customer Phone",
            "Customer Name",
            "Customer Address",
            "Customer Locality",
            "Persons",
            "Cancel Reason",
            "Total Amount",
            "Total Tax",
            "Discount Amount",
            "Delivery Charge",
            "Container Charge",
            "Service Charge",
            "Additional Charge",
            "Waived Off",
            "Round Off",
            "Total",
        ]
        ws.append(headers)

        # ---------------- DATA ----------------
        async for doc in cursor:

            # ✅ Date format
            inv_date = doc.get("invoiceDateTime", "")
            if isinstance(inv_date, datetime):
                inv_date = inv_date.strftime("%d-%m-%Y %H:%M")

            # ✅ Payment type fix (list → string)
            pay_type = doc.get("paymentType") or ""
            if isinstance(pay_type, list):
                pay_type = ", ".join(map(str, pay_type))

            row = [
                doc.get("branchName", ""),
                doc.get("invoiceNo", ""),
                doc.get("gstNo", ""),
                inv_date,
                doc.get("kotNo", ""),
                pay_type,
                doc.get("paymentDescription", ""),
                doc.get("salesType", ""),
                doc.get("status", ""),
                doc.get("subOrderType", ""),
                doc.get("area", ""),
                doc.get("virtualBranch", ""),
                doc.get("assign_to", ""),
                doc.get("groupName", ""),
                doc.get("customerPhoneNumber", ""),
                doc.get("customerName", ""),
                doc.get("customerAddress", ""),
                doc.get("customerLocality", ""),
                doc.get("persons", ""),
                doc.get("cancelReason", ""),
                doc.get("totalAmount", ""),
                doc.get("totalTax", ""),
                doc.get("discountAmount", ""),
                doc.get("deliveryCharge", ""),
                doc.get("containerCharge", ""),
                doc.get("serviceCharge", ""),
                doc.get("additionalCharge", ""),
                doc.get("waivedOff", ""),
                doc.get("roundOff", ""),
                doc.get("total", ""),
            ]

            ws.append(row)

        wb.save(output)
        output.seek(0)

        filename = f"PayMode_YenERP_{datetime.now().strftime('%d-%m-%Y_%H-%M')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
