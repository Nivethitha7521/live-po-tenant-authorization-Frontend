import tempfile
from math import ceil
from typing import List, Optional, Dict, Any

from fastapi import APIRouter, Query
from fastapi.responses import FileResponse
from openpyxl import Workbook
from datetime import datetime, timedelta, timezone
from fastapi import Depends
from dependencies.auth import validate_token
from middlewares.permission_middleware import check_permission
from excel import parse_date
from globalsReport.allfuntions import get_qty_by_uom, to_int
from .models import (
    DispatchDateDropdownResponse,
    VarianceDropdownResponse,
    branchDropdownResponse,
    driverDropdownResponse,
)
from db.collections import (
    dispatch,
    ItemMaster,
)

router = APIRouter()

# --- CONSTANTS ---

REPORT_HEADERS_DISPATCH = [
    "DocNo",
    "dispatchNo",
    "LineID",
    "ItemCode",
    "ItemName",
    "Group",
    "Sub-Group",
    "UOM",
    "HSN",
    "Qty",
    "Price",
    "Total",
    "TaxCode",
    "TaxAmt",
    "LoginID",
    "LoginName",
    "LastName",
    "LocationId",
    "Location",
    "VehicleName",
    "VehicleNo",
    "Driver-ID",
    "DriverName",
    "Initial",
    "Date",
    "DespTime",
    "LeadTime",
    "ExpDate",
]

REPORT_HEADERS_RECEIVE = [
    "DocNo",
    "LineID",
    "ItemCode",
    "ItemName",
    "Group",
    "Sub-Group",
    "UOM",
    "HSN",
    "ReceivedQty",
    "Price",
    "Total",
    "TaxCode",
    "TaxAmt",
    "Date",
    "Receive.Time",
    "DriverCode",
    "DriverName",
    "VehicleNo",
    "LoginID",
    "LoginName",
    "Loc.ID",
    "Location",
    "DespatchNo",
]

REPORT_HEADERS_LOCATION_RECEIVE = [
    "DocNo",
    "LineID",
    "ItemCode",
    "ItemName",
    "Group",
    "Sub-Group",
    "UOM",
    "HSN",
    "ReceivedQty",
    "Price",
    "Total",
    "TaxCode",
    "TaxAmt",
    "Date",
    "Receive.Time",
    "DriverCode",
    "DriverName",
    "VehicleNo",
    "LoginID",
    "LoginName",
    "Loc.ID",
    "Location",
    "DespatchNo",
]

# --- HELPERS ---


def safe_index(field_list: Any, idx: int) -> Any:
    if isinstance(field_list, list) and len(field_list) > idx:
        return field_list[idx]
    return None


def process_line_item(
    d: Dict, idx: int, item_info: Dict, status: str = "dispatched"
) -> Dict:
    # FIX: After $unwind, 'varianceName' is a STRING (the item name itself), not an array.
    # We do NOT use safe_index for varianceName anymore.
    v_name = d.get("varianceName")
    print(v_name)

    # These fields remain ARRAYS in the document, so we use safe_index
    qty = safe_index(d.get("qty"), idx) or 0
    price = safe_index(d.get("price"), idx) or 0
    total_amt = safe_index(d.get("amount"), idx) or (qty * price)

    cat_name = safe_index(d.get("categoryName"), idx) or item_info.get("category", "")
    subcat_name = safe_index(d.get("subCategoryName"), idx) or item_info.get(
        "subCategory", ""
    )

    if status == "dispatched":
        dispatch_date = d.get("date")
        date_str, time_str, exp_date = None, None, None
        lead_time = item_info.get("shelfLife")

        if dispatch_date:
            date_str = dispatch_date.strftime("%d-%m-%Y")
            time_str = dispatch_date.strftime("%H:%M")
            if lead_time:
                try:
                    exp_date = (
                        dispatch_date + timedelta(days=int(lead_time))
                    ).strftime("%d-%m-%Y")
                except ValueError:
                    pass

        return {
            "DocNo": str(d.get("_id", "")),
            "LineID": idx + 1,
            "ItemCode": safe_index(d.get("itemCode"), idx),
            "ItemName": v_name,  # This is now correct
            "Group": cat_name,
            "Sub-Group": subcat_name,
            "UOM": safe_index(d.get("uom"), idx),
            "HSN": item_info.get("hsnCode"),
            "Qty": qty,
            "Price": price,
            "Total": total_amt,
            "TaxCode": item_info.get("TaxCode"),
            "TaxAmt": None,
            "LoginID": d.get("loginId"),
            "LoginName": d.get("createdBy"),
            "LastName": d.get("lastName"),
            "LocationId": d.get("locationId"),
            "Location": d.get("branchName"),
            "VehicleName": d.get("vehicleName") or "NA",
            "VehicleNo": d.get("vehicleNumber"),
            "Driver-ID": d.get("driverId"),
            "DriverName": d.get("driverFirstName"),
            "Initial": d.get("Initial"),
            "Date": date_str,
            "DespTime": time_str,
            "LeadTime": lead_time,
            "ExpDate": exp_date,
        }
    else:
        rt = d.get("receivedTime")
        # Check if itemName array exists, otherwise use v_name
        item_name_arr = safe_index(d.get("varianceName"), idx)

        return {
            "DocNo": str(d.get("_id")),
            "LineID": idx + 1,
            "ItemCode": safe_index(d.get("itemCode"), idx),
            "ItemName": item_name_arr or v_name,
            "Group": item_info.get("category"),
            "Sub-Group": item_info.get("subCategory"),
            "UOM": safe_index(d.get("uom"), idx),
            "HSN": str(item_info.get("hsnCode")) if item_info.get("hsnCode") else None,
            "ReceivedQty": qty,
            "Price": price,
            "Total": total_amt,
            "TaxCode": item_info.get("TaxCode"),
            "TaxAmt": None,
            "Date": rt.strftime("%d-%m-%Y") if rt else None,
            "Receive.Time": rt.strftime("%H:%M") if rt else None,
            "LoginID": d.get("loginId"),
            "LoginName": d.get("createdBy"),
            "Loc.ID": d.get("locationId"),
            "Location": d.get("branchName"),
            "VehicleNo": d.get("vehicleNumber"),
            "DriverCode": d.get("Driver-ID") or d.get("driverId"),
            "DriverName": d.get("driverFirstName"),
            "DespatchNo": d.get("dispatchNo"),
        }


# --- DROPDOWNS ---


async def _get_date_dropdown_impl(status: str):
    pipeline = [
        {"$match": {"date": {"$type": "date"}, "status": status}},
        {
            "$group": {
                "_id": None,
                "years": {"$addToSet": {"$year": "$date"}},
                "months": {"$addToSet": {"$month": "$date"}},
                "days": {"$addToSet": {"$dayOfMonth": "$date"}},
            }
        },
    ]
    result = await dispatch.aggregate(pipeline).to_list(1)
    if not result:
        return DispatchDateDropdownResponse(yearIn=[], monthIn=[], daysIn=[])
    return DispatchDateDropdownResponse(
        yearIn=sorted(map(str, result[0]["years"])),
        monthIn=sorted(f"{m:02d}" for m in result[0]["months"]),
        daysIn=sorted(result[0]["days"]),
    )


@router.get("/date-dropdown", response_model=DispatchDateDropdownResponse)
async def get_dispatch_date_dropdown(user=Depends(validate_token),
    permissions=Depends(check_permission("yenerp", "posreport", "read"))):
    return await _get_date_dropdown_impl("dispatched")


@router.get("/receive/date-dropdown", response_model=DispatchDateDropdownResponse)
async def get_receive_date_dropdown(user=Depends(validate_token),
    permissions=Depends(check_permission("yenerp", "posreport", "read"))):
    return await _get_date_dropdown_impl("received")


# --- REPORT & EXPORT LOGIC ---


async def _generate_report_data(
    status: str,
    page: int,
    limit: int,
    startDate: Optional[str],
    endDate: Optional[str],
    varianceName: Optional[List[str]],
    branchName: Optional[List[str]],
    driverName: Optional[List[str]],
):
    query = {"status": status}

    # Date Filters
    if startDate or endDate:
        date_filter = {}
        if startDate:
            start = (
                parse_date(startDate)
                if status == "received"
                else datetime.fromisoformat(startDate)
            )
            s_date = datetime.combine(start.date(), datetime.min.time())
            if status == "received":
                s_date = s_date.replace(tzinfo=timezone.utc)
            date_filter["$gte"] = s_date

        if endDate:
            end = (
                parse_date(endDate)
                if status == "received"
                else datetime.fromisoformat(endDate)
            )
            e_date = datetime.combine(end.date(), datetime.max.time())
            if status == "received":
                e_date = e_date.replace(tzinfo=timezone.utc)
            date_filter["$lte"] = e_date
        query["date"] = date_filter

    if branchName:
        query["locationId"] = {"$in": branchName}
    if driverName:
        query["driverFirstName"] = {"$in": driverName}

    pipeline = [
        {"$match": query},
        {"$unwind": {"path": "$varianceName", "includeArrayIndex": "idx"}},
    ]

    # Filter by itemCode (varianceName param holds Item Codes)
    if varianceName:
        pipeline.append(
            {
                "$match": {
                    "$expr": {
                        "$in": [{"$arrayElemAt": ["$itemCode", "$idx"]}, varianceName]
                    }
                }
            }
        )

    pipeline.extend(
        [
            {
                "$lookup": {
                    "from": "ItemMaster",
                    "localField": "varianceName",
                    "foreignField": "varianceName",
                    "as": "item_join",
                }
            },
            {"$set": {"item_join": {"$arrayElemAt": ["$item_join", 0]}}},
            {
                "$facet": {
                    "metadata": [{"$count": "total"}],
                    "data": [{"$skip": (page - 1) * limit}, {"$limit": limit}],
                }
            },
        ]
    )

    result = await dispatch.aggregate(pipeline, allowDiskUse=True).to_list(1)
    total_items = (
        result[0]["metadata"][0]["total"] if result and result[0]["metadata"] else 0
    )
    total_pages = ceil(total_items / limit) if total_items > 0 else 0

    entries = []
    for d in result[0]["data"]:
        item_info = d.get("item_join") or {}
        entry_dict = process_line_item(d, d["idx"], item_info, status=status)
        entries.append(entry_dict)

    return {
        "items": entries,
        "page": page,
        "limit": limit,
        "total": total_items,
        "totalPages": total_pages,
    }


async def _generate_excel_file(
    status: str,
    startDate: Optional[str],
    endDate: Optional[str],
    varianceName: Optional[List[str]],
    branchName: Optional[List[str]],
    driverName: Optional[List[str]],
):
    query = {"status": status}
    if startDate or endDate:
        date_filter = {}
        if startDate:
            start = (
                parse_date(startDate)
                if status == "received"
                else datetime.fromisoformat(startDate)
            )
            s_date = datetime.combine(start.date(), datetime.min.time())
            if status == "received":
                s_date = s_date.replace(tzinfo=timezone.utc)
            date_filter["$gte"] = s_date
        if endDate:
            end = (
                parse_date(endDate)
                if status == "received"
                else datetime.fromisoformat(endDate)
            )
            e_date = datetime.combine(end.date(), datetime.max.time())
            if status == "received":
                e_date = e_date.replace(tzinfo=timezone.utc)
            date_filter["$lte"] = e_date
        query["date"] = date_filter
    if branchName:
        query["locationId"] = {"$in": branchName}
    if driverName:
        query["driverFirstName"] = {"$in": driverName}

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Report")
    headers = (
        REPORT_HEADERS_RECEIVE if status == "received" else REPORT_HEADERS_DISPATCH
    )
    ws.append(headers)

    pipeline = [
        {"$match": query},
        {"$unwind": {"path": "$varianceName", "includeArrayIndex": "idx"}},
    ]

    if varianceName:
        pipeline.append(
            {
                "$match": {
                    "$expr": {
                        "$in": [{"$arrayElemAt": ["$itemCode", "$idx"]}, varianceName]
                    }
                }
            }
        )


    pipeline.extend(
        [
            {
                "$lookup": {
                    "from": "ItemMaster",
                    "localField": "varianceName",
                    "foreignField": "varianceName",
                    "as": "item_info",
                }
            },
            {"$set": {"item_info": {"$arrayElemAt": ["$item_info", 0]}}},
        ]
    )

    cursor = dispatch.aggregate(pipeline, allowDiskUse=True)

    async for d in cursor:
        item_info = d.get("item_info") or {}
        entry_dict = process_line_item(d, d["idx"], item_info, status=status)
        row = [entry_dict.get(h) for h in headers]
        ws.append(row)

    filename = f"{'Receive' if status == 'received' else 'Dispatch'}_YenERP_{datetime.now().strftime('%d-%m-%Y_%H-%M')}.xlsx"
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        return tmp.name, filename


# --- ENDPOINTS ---


@router.get("/report", response_model=dict)
async def get_dispatch_report(
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1),
    startDate: Optional[str] = Query(None),
    endDate: Optional[str] = Query(None),
    varianceName: Optional[List[str]] = Query(None),
    branchName: Optional[List[str]] = Query(None),
    driverName: Optional[List[str]] = Query(None),
    user=Depends(validate_token),
    permissions=Depends(check_permission("yenerp", "posreport", "read"))
):
    return await _generate_report_data(
        "dispatched",
        page,
        limit,
        startDate,
        endDate,
        varianceName,
        branchName,
        driverName,
    )


@router.get("/export")
async def export_dispatch_excel(
    startDate: Optional[str] = Query(None),
    endDate: Optional[str] = Query(None),
    varianceName: Optional[List[str]] = Query(None),
    branchName: Optional[List[str]] = Query(None),
    driverName: Optional[List[str]] = Query(None),
    user=Depends(validate_token),
    permissions=Depends(check_permission("yenerp", "posreport", "read"))
):
    tmp_path, filename = await _generate_excel_file(
        "dispatched", startDate, endDate, varianceName, branchName, driverName
    )
    return FileResponse(
        tmp_path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/receive/report", response_model=dict)
async def get_receive_report(
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1),
    startDate: Optional[str] = Query(None),
    endDate: Optional[str] = Query(None),
    varianceName: Optional[List[str]] = Query(None),
    branchName: Optional[List[str]] = Query(None),
    driverName: Optional[List[str]] = Query(None),
    user=Depends(validate_token),
    permissions=Depends(check_permission("yenerp", "posreport", "read"))
):
    return await _generate_report_data(
        "received",
        page,
        limit,
        startDate,
        endDate,
        varianceName,
        branchName,
        driverName,
    )


@router.get("/receive/export")
async def export_receive_excel(
    startDate: Optional[str] = Query(None),
    endDate: Optional[str] = Query(None),
    varianceName: Optional[List[str]] = Query(None),
    branchName: Optional[List[str]] = Query(None),
    driverName: Optional[List[str]] = Query(None),
    user=Depends(validate_token),
    permissions=Depends(check_permission("yenerp", "posreport", "read"))
):
    tmp_path, filename = await _generate_excel_file(
        "received", startDate, endDate, varianceName, branchName, driverName
    )
    return FileResponse(
        tmp_path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# --- LOCATION RECEIVE ---


@router.get("/locationreceive/report", response_model=dict)
async def get_location_receive_report(
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1),
    startDate: Optional[str] = Query(None),
    endDate: Optional[str] = Query(None),
    varianceName: Optional[List[str]] = Query(None),
    branchName: Optional[List[str]] = Query(None),
    driverName: Optional[List[str]] = Query(None),
    user=Depends(validate_token),
    permissions=Depends(check_permission("yenerp", "posreport", "read"))
):
    match_query = {"status": "received"}

    if startDate or endDate:
        date_filter = {}
        if startDate:
            start = parse_date(startDate)
            date_filter["$gte"] = datetime.combine(
                start.date(), datetime.min.time()
            ).replace(tzinfo=timezone.utc)
        if endDate:
            end = parse_date(endDate)
            date_filter["$lte"] = datetime.combine(
                end.date(), datetime.max.time()
            ).replace(tzinfo=timezone.utc)
        match_query["date"] = date_filter

    if branchName:
        match_query["branchName"] = {"$in": branchName}
    if driverName:
        match_query["driverFirstName"] = {"$in": driverName}

    pipeline = [
        {"$match": match_query},
        {"$unwind": {"path": "$varianceName", "includeArrayIndex": "idx"}},
    ]

    if varianceName:
        pipeline.append(
            {
                "$match": {
                    "$expr": {
                        "$in": [{"$arrayElemAt": ["$itemCode", "$idx"]}, varianceName]
                    }
                }
            }
        )


    pipeline.append(
        {
            "$facet": {
                "metadata": [{"$count": "total"}],
                "data": [{"$skip": (page - 1) * limit}, {"$limit": limit}],
            }
        }
    )

    results = await dispatch.aggregate(pipeline).to_list(length=1)
    facet_result = results[0] if results else {"metadata": [], "data": []}
    total = facet_result["metadata"][0]["total"] if facet_result["metadata"] else 0
    raw_data = facet_result["data"]

    final_entries = []
    for doc in raw_data:
        # FIX: Use direct get for varianceName after unwind
        v_name = doc.get("varianceName")
        idx = doc.get("idx")

        item_info = await ItemMaster.find_one(
            {"varianceName": {"$regex": f"^{v_name}$", "$options": "i"}}
        )
        qty, uom = get_qty_by_uom(doc, idx)
        price = safe_index(doc.get("price"), idx) or 0
        amount = safe_index(doc.get("amount"), idx) or 0
        receive_date = doc.get("receivedTime")
        dispatch_date = doc.get("date")

        entry = {
            "DocNo": str(doc.get("_id"))[-5:],
            "dispatchNo": doc.get("dispatchNo"),
            "LineID": idx + 1,
            "ItemCode": safe_index(doc.get("itemCode"), idx),
            "ItemName": safe_index(doc.get("varianceName"), idx) or v_name,
            "Group": safe_index(doc.get("categoryName"), idx),
            "Sub-Group": safe_index(doc.get("subCategoryName"), idx),
            "UOM": uom,
            "HSN": (
                str(item_info.get("hsnCode"))
                if item_info and item_info.get("hsnCode")
                else None
            ),
            "ReceivedQty": qty,
            "Price": price,
            "Total": amount,
            "TaxCode": item_info.get("TaxCode") if item_info else None,
            "TaxAmt": None,
            "Date": dispatch_date.strftime("%d-%m-%Y") if dispatch_date else None,
            "ReceiveTime": receive_date.strftime("%H:%M") if receive_date else None,
            "LoginID": doc.get("loginId"),
            "LoginName": doc.get("createdBy"),
            "Loc.ID": doc.get("locationId"),
            "Location": doc.get("branchName"),
            "VehicleNo": doc.get("vehicleNumber"),
            "DriverCode": doc.get("Driver-ID") or doc.get("driverId"),
            "DriverName": doc.get("driverFirstName"),
            "DespatchNo": doc.get("dispatchNo"),
        }
        final_entries.append(entry)

    return {
        "items": final_entries,
        "page": page,
        "limit": limit,
        "total": total,
        "totalPages": ceil(total / limit),
    }


@router.get("/locationreceive/export")
async def export_location_receive_excel(
    startDate: Optional[str] = Query(None),
    endDate: Optional[str] = Query(None),
    varianceName: Optional[List[str]] = Query(None),
    branchName: Optional[List[str]] = Query(None),
    driverName: Optional[List[str]] = Query(None),
    user=Depends(validate_token),
    permissions=Depends(check_permission("yenerp", "posreport", "read"))
):
    match_query = {"status": "received"}

    if startDate or endDate:
        date_filter = {}
        if startDate:
            start = parse_date(startDate)
            date_filter["$gte"] = datetime.combine(
                start.date(), datetime.min.time()
            ).replace(tzinfo=timezone.utc)
        if endDate:
            end = parse_date(endDate)
            date_filter["$lte"] = datetime.combine(
                end.date(), datetime.max.time()
            ).replace(tzinfo=timezone.utc)
        match_query["date"] = date_filter

    if branchName:
        match_query["branchName"] = {"$in": branchName}
    if driverName:
        match_query["driverFirstName"] = {"$in": driverName}

    pipeline = [
        {"$match": match_query},
        {"$unwind": {"path": "$varianceName", "includeArrayIndex": "idx"}},
    ]

    if varianceName:
        pipeline.append(
            {
                "$match": {
                    "$expr": {
                        "$in": [{"$arrayElemAt": ["$itemCode", "$idx"]}, varianceName]
                    }
                }
            }
        )


    cursor = dispatch.aggregate(pipeline, allowDiskUse=True)

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("LocationReceive")
    headers = REPORT_HEADERS_LOCATION_RECEIVE
    ws.append(headers)

    async for doc in cursor:
        # FIX: Use direct get for varianceName after unwind
        v_name = doc.get("varianceName")
        idx = doc.get("idx")

        item_info = await ItemMaster.find_one(
            {"varianceName": {"$regex": f"^{v_name}$", "$options": "i"}}
        )
        dispatch_date = doc.get("date")
        receive_date = doc.get("receivedTime")
        qty, uom = get_qty_by_uom(doc, idx)
        price = safe_index(doc.get("price"), idx) or 0
        amount = safe_index(doc.get("amount"), idx) or 0

        row_data = {
            "DocNo": str(doc.get("_id")),
            "LineID": idx + 1,
            "ItemCode": safe_index(doc.get("itemCode"), idx),
            "ItemName": safe_index(doc.get("varianceName"), idx) or v_name,
            "Group": safe_index(doc.get("categoryName"), idx),
            "Sub-Group": safe_index(doc.get("subCategoryName"), idx),
            "UOM": uom,
            "HSN": (
                str(item_info.get("hsnCode"))
                if item_info and item_info.get("hsnCode")
                else None
            ),
            "ReceivedQty": qty,
            "Price": price,
            "Total": amount,
            "TaxCode": item_info.get("TaxCode") if item_info else None,
            "TaxAmt": "",
            "Date": dispatch_date.strftime("%d-%m-%Y") if dispatch_date else None,
            "Receive.Time": receive_date.strftime("%H:%M") if receive_date else None,
            "DriverCode": doc.get("Driver-ID") or doc.get("driverId"),
            "DriverName": doc.get("driverFirstName"),
            "VehicleNo": doc.get("vehicleNumber"),
            "LoginID": to_int(doc.get("loginId")),
            "LoginName": doc.get("createdBy"),
            "Loc.ID": doc.get("locationId"),
            "Location": doc.get("branchName"),
            "DespatchNo": doc.get("dispatchNo"),
        }
        row = [row_data.get(h) for h in headers]
        ws.append(row)

    filename = (
        f"DispatchLocReceive_YenERP_{datetime.now().strftime('%d-%m-%Y_%H-%M')}.xlsx"
    )
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        return FileResponse(
            tmp.name,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
