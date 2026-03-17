import csv
import io
import traceback
from fastapi import APIRouter, Body, HTTPException, Query,Request,Depends
from typing import List, Optional, Dict
from datetime import datetime, timedelta, date
from dependencies.auth import validate_token
from middlewares.permission_middleware import check_permission
import asyncio
import math

from fastapi.responses import StreamingResponse
from openpyxl.styles import PatternFill, Font

from openpyxl import Workbook

from OutletInventory.funtions import (
    build_category_map,
    build_mongo_filter_from_params,
    build_subcategory_map,
    get_filter_field_options,
    ist_now,
    ist_yesterday,
)
from OutletInventory.routes import master_collection
from OutletInventoryVariance.functions import (
    calculate_quantities,
    daterange,
    get_dispatch_mapping_agg,
    get_sales_mapping_agg,
    get_sales_return_mapping_agg,
    get_stock_transfer_mapping_agg,
    get_warehousereturn_mapping_agg,
    get_wastagereturn_mapping_agg,
    match_dispatch,
    match_sales,
    match_sales_return,
    match_stock_transfer,
    match_warehouseReturn,
    match_wastage,
    normalize_key,
    safe_float,
)
from OutletInventoryVariance.models import ApproveItemRequest, ItemResponse

from db.collections import (
    approvedstocks_fg_collection,
    closingstocks_collection,
    inventory_stock_collection,
    location_collection,
    uom_collection,
)
from funtions import force_float_conversion, round_all_numbers

router = APIRouter()


# ==================== UOM PRECISION HELPER ====================


async def get_uom_precision_map() -> Dict[str, int]:
    """
    Fetches all UOM documents and returns a map of {uomId: precision}.
    Default precision is 0 if not specified.
    """
    cursor = uom_collection().find({}, {"uomId": 1, "precision": 1})
    docs = await cursor.to_list(None)

    return {doc["uomId"]: doc.get("precision", 0) for doc in docs if doc.get("uomId")}


def round_by_precision(value: float, precision: int) -> float:
    """
    Rounds a value based on precision.
    If precision is 0, returns an integer (or float with .0).
    """
    if value is None:
        return 0.0

    # Python's round handles precision directly
    # e.g., round(10.567, 2) -> 10.57
    # e.g., round(10.567, 0) -> 11.0

    # If you want strict integer for precision 0, uncomment next line
    # return int(round(value, precision)) if precision == 0 else round(value, precision)

    return round(value, precision)


# ==================== OPENING STOCK HELPER ====================


async def get_opening_stocks_from_inventory_snapshot(tenant_id:str,
    location_id: str,
    query_date: Optional[date] = None,
):
    try:
        inventory_coll = closingstocks_collection(tenant_id)

        if query_date is None:
            query_date = date.today()

        snapshot_date = query_date - timedelta(days=1)
        snapshot_date_str = snapshot_date.strftime("%Y-%m-%d")

        #  Fetch yesterday snapshot rows
        cursor = inventory_coll.find(
            {
                "locationId": location_id,
                "date": snapshot_date_str,
            },
            {
                "itemCode": 1,
                "systemStock": 1,
            },
        )

        docs = await cursor.to_list(None)

        opening_map = {
            doc["itemCode"]: safe_float(doc.get("systemStock", 0))
            for doc in docs
            if doc.get("itemCode")
        }

        return {
            "opening_map": opening_map,
            "snapshot_date": snapshot_date,
        }

    except Exception:
        traceback.print_exc()
        return {"opening_map": {}, "snapshot_date": None}


# ==================== MAIN GET ENDPOINT ====================
@router.get("/")
async def get_items(request:Request,
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1, le=100),
    branch: str = Query(...),
    category: Optional[str] = Query(None),
    subCategory: Optional[str] = Query(None),
    itemName: Optional[str] = Query(None),
    varianceName: Optional[str] = Query(None),
    queryDate: Optional[date] = Query(None),
    include_filter_options: bool = Query(True),
    only_filter_options: bool = Query(False),
    categoryPage: int = Query(1, ge=1),
    categoryLimit: int = Query(20, ge=1),
    categorySearch: Optional[str] = Query(None),
    subCategoryPage: int = Query(1, ge=1),
    subCategoryLimit: int = Query(20, ge=1),
    subCategorySearch: Optional[str] = Query(None),
    itemNamePage: int = Query(1, ge=1),
    itemNameLimit: int = Query(20, ge=1),
    itemNameSearch: Optional[str] = Query(None),
    varianceNamePage: int = Query(1, ge=1),
    varianceNameLimit: int = Query(20, ge=1),
    varianceNameSearch: Optional[str] = Query(None),
    user = Depends(validate_token),
    permissions: dict = Depends(check_permission("yenerp","physicalstockvariancemodification","read")),
):
    tenant_id = request.state.tenant_id
    # ---------------- Build filters ---------------- #
    filters = build_mongo_filter_from_params(
        category=category,
        subCategory=subCategory,
        itemName=itemName,
        varianceName=varianceName,
    )

    skip = (page - 1) * limit

    col = master_collection()

    category_map, subcategory_map, uom_precision_map = await asyncio.gather(
        build_category_map(),
        build_subcategory_map(),
        get_uom_precision_map(),  # Fetch UOM precision map
    )

    # ---------------- ONLY FILTER OPTIONS ---------------- #
    if only_filter_options:
        response = {}

        if include_filter_options:
            filter_tasks = [
                get_filter_field_options(
                    collection=col,
                    field="category",
                    page=categoryPage,
                    limit=categoryLimit,
                    category=category,
                    subCategory=subCategory,
                    itemName=itemName,
                    varianceName=varianceName,
                    search_filter=categorySearch,
                    category_map=category_map,
                    subcategory_map=subcategory_map,
                ),
                get_filter_field_options(
                    collection=col,
                    field="subCategory",
                    page=subCategoryPage,
                    limit=subCategoryLimit,
                    category=category,
                    subCategory=subCategory,
                    itemName=itemName,
                    varianceName=varianceName,
                    search_filter=subCategorySearch,
                    category_map=category_map,
                    subcategory_map=subcategory_map,
                ),
                get_filter_field_options(
                    collection=col,
                    field="itemName",
                    page=itemNamePage,
                    limit=itemNameLimit,
                    category=category,
                    subCategory=subCategory,
                    itemName=itemName,
                    varianceName=varianceName,
                    search_filter=itemNameSearch,
                    category_map=category_map,
                    subcategory_map=subcategory_map,
                ),
                get_filter_field_options(
                    collection=col,
                    field="varianceName",
                    page=varianceNamePage,
                    limit=varianceNameLimit,
                    category=category,
                    subCategory=subCategory,
                    itemName=itemName,
                    varianceName=varianceName,
                    search_filter=varianceNameSearch,
                    category_map=category_map,
                    subcategory_map=subcategory_map,
                ),
            ]

            filter_results = await asyncio.gather(*filter_tasks)
            response["filterOptions"] = {
                "category": filter_results[0],
                "subCategory": filter_results[1],
                "itemName": filter_results[2],
                "varianceName": filter_results[3],
            }

        return response

    # ---------------- Parallel fetch ---------------- #
    total_task = col.count_documents(filters)
    items_task = col.find(filters).skip(skip).limit(limit).to_list(length=limit)

    # Fetch all mappings
    dispatch_task = get_dispatch_mapping_agg(branch, queryDate)
    sales_task = get_sales_mapping_agg(branch, queryDate)
    sales_return_task = get_sales_return_mapping_agg(branch, queryDate)
    warehouse_task = get_warehousereturn_mapping_agg(branch, queryDate)
    wastage_task = get_wastagereturn_mapping_agg(branch, queryDate)
    transfer_task = get_stock_transfer_mapping_agg(branch, queryDate)

    # Fetch opening stocks from yesterday
    opening_stock_task = get_opening_stocks_from_inventory_snapshot(tenant_id=tenant_id,
        location_id=branch,
        query_date=queryDate,
    )

    # Execute all tasks concurrently
    results = await asyncio.gather(
        total_task,
        items_task,
        dispatch_task,
        sales_task,
        sales_return_task,
        warehouse_task,
        wastage_task,
        transfer_task,
        opening_stock_task,
    )

    (
        total,
        raw_items,
        dispatch_map,
        sales_data,
        sales_return_data,
        warehouse_map,
        wastage_data,
        transfer_data,
        yesterday_opening_map,
    ) = results

    # ---------------- Fetch inventory stocks in bulk ---------------- #
    item_codes = [item.get("itemCode") for item in raw_items if item.get("itemCode")]

    inventory_docs = (
        await inventory_stock_collection(tenant_id)
        .find({"locationId": branch, "itemCode": {"$in": item_codes}})
        .to_list(length=len(item_codes))
    )

    inventory_map = {
        f"{doc['itemCode']}__{doc['locationId']}": doc for doc in inventory_docs
    }

    # ---------------- Process items ---------------- #
    items = []
    for item in raw_items:
        variance = item.get("varianceName")
        key = normalize_key(variance)

        # Get UOM ID and Precision
        uom_id = item.get("item_Uom")
        precision = uom_precision_map.get(uom_id, 0)  # Default to 0 if UOM not found

        # Opening stock from yesterday snapshot
        opening_stock = safe_float(
            yesterday_opening_map["opening_map"].get(item.get("itemCode"), 0.0)
        )

        # Apply precision to opening stock immediately
        opening_stock = round_by_precision(opening_stock, precision)

        category_id = item.get("category")
        subcategory_id = item.get("subCategory")
        item_code = item.get("itemCode")
        variance_name = item.get("varianceName")

        # Base item data
        result = {
            "id": str(item["_id"]),
            "itemCode": item_code,
            "uomPrecision": precision,  # Include precision in response for frontend
            "category": {
                "id": category_id,
                "name": category_map.get(category_id),
            },
            "subCategory": {
                "id": subcategory_id,
                "name": subcategory_map.get(subcategory_id),
            },
            "itemName": {
                "id": subcategory_id,
                "name": subcategory_map.get(subcategory_id),
            },
            "varianceName": {
                "id": item_code,
                "name": variance_name,
            },
            "openingStockQty": opening_stock,
            "receivedQty": 0.0,
            "stockTransferInQty": 0.0,
            "stockTransferOutQty": 0.0,
            "salesQty": 0.0,
            "salesReturnQty": 0.0,
            "wastageReceivedQty": 0.0,
            "wastageReturnQty": 0.0,
            "closingStockQty": 0.0,
        }

        # Apply today's movements
        result = match_dispatch(result, dispatch_map)
        result = match_sales(result, sales_data["full_map"])
        result = match_sales_return(result, sales_return_data["full_map"])
        result = match_warehouseReturn(result, warehouse_map["full_map"])
        result = match_wastage(result, wastage_data["full_map"])
        result = match_stock_transfer(result, transfer_data["full_map"])

        # Calculate current system quantity based on opening + movements
        # Note: Ensure your calculate_quantities uses the precision if modified internally,
        # otherwise we round the final output below.
        result = calculate_quantities(result)

        # ---------------- GET STOCK FROM INVENTORY ---------------- #
        inventory_key = f"{item.get('itemCode')}__{branch}"
        inventory_doc = inventory_map.get(inventory_key)

        if inventory_doc:
            system_stock = safe_float(inventory_doc.get("systemStock", 0))
            physical_stock = safe_float(inventory_doc.get("physicalStock", 0))
            stored_variance = safe_float(inventory_doc.get("variance", 0))
            status = inventory_doc.get("status", "pending")
            updated_at = inventory_doc.get("updatedAt")

            variance_val = safe_float(inventory_doc.get("variance", 0))

            updated_today = False
            if queryDate and updated_at:
                updated_today = updated_at.date() == queryDate
            elif not queryDate and updated_at:
                updated_today = True

            if not updated_today:
                physical_stock_display = "-"
                variance_display = "-"
                show_approve_button = False
                status_display = "Not Available"
            else:
                # Apply precision to physical stock and variance
                physical_stock_display = round_by_precision(physical_stock, precision)
                variance_display = round_by_precision(variance_val, precision)
                show_approve_button = variance_val != 0 and status != "approved"
                status_display = status.capitalize() if status else "Pending"

            result["updatedCurrentSystemQty"] = round_by_precision(
                system_stock, precision
            )
            result["physicalClosingQty"] = physical_stock_display
            result["stockVariance"] = variance_display
            result["approvalStatus"] = status_display
            result["approveButton"] = show_approve_button
            result["lastUpdated"] = (
                updated_at.strftime("%Y-%m-%d %H:%M:%S") if updated_at else "-"
            )
        else:
            result["updatedCurrentSystemQty"] = 0.0
            result["physicalClosingQty"] = "-"
            result["stockVariance"] = "-"
            result["approvalStatus"] = "Not Available"
            result["approveButton"] = False
            result["lastUpdated"] = "-"

        # ---------------- FINAL ROUNDING BY UOM PRECISION ---------------- #
        # Ensure all quantity fields respect the UOM precision
        qty_fields = [
            "openingStockQty",
            "receivedQty",
            "stockTransferInQty",
            "stockTransferOutQty",
            "salesQty",
            "salesReturnQty",
            "wastageReceivedQty",
            "wastageReturnQty",
            "closingStockQty",
            "updatedCurrentSystemQty",
        ]

        for field in qty_fields:
            if field in result and isinstance(result[field], (int, float)):
                result[field] = round_by_precision(result[field], precision)

        items.append(result)

    # ---------------- Build response ---------------- #
    response = {
        "filteredItems": {
            "total": total,
            "page": page,
            "limit": limit,
            "count": len(items),
            "items": items,
        }
    }

    # Include filter options if requested
    if include_filter_options:
        filter_tasks = [
            get_filter_field_options(
                collection=col,
                field="category",
                page=categoryPage,
                limit=categoryLimit,
                category=category,
                subCategory=subCategory,
                itemName=itemName,
                varianceName=varianceName,
                search_filter=categorySearch,
                category_map=category_map,
                subcategory_map=subcategory_map,
            ),
            get_filter_field_options(
                collection=col,
                field="subCategory",
                page=subCategoryPage,
                limit=subCategoryLimit,
                category=category,
                subCategory=subCategory,
                itemName=itemName,
                varianceName=varianceName,
                search_filter=subCategorySearch,
                category_map=category_map,
                subcategory_map=subcategory_map,
            ),
            get_filter_field_options(
                collection=col,
                field="itemName",
                page=itemNamePage,
                limit=itemNameLimit,
                category=category,
                subCategory=subCategory,
                itemName=itemName,
                varianceName=varianceName,
                search_filter=itemNameSearch,
                category_map=category_map,
                subcategory_map=subcategory_map,
            ),
            get_filter_field_options(
                collection=col,
                field="varianceName",
                page=varianceNamePage,
                limit=varianceNameLimit,
                category=category,
                subCategory=subCategory,
                itemName=itemName,
                varianceName=varianceName,
                search_filter=varianceNameSearch,
                category_map=category_map,
                subcategory_map=subcategory_map,
            ),
        ]

        filter_results = await asyncio.gather(*filter_tasks)
        response["filterOptions"] = {
            "category": filter_results[0],
            "subCategory": filter_results[1],
            "itemName": filter_results[2],
            "varianceName": filter_results[3],
        }

    # We skip round_all_numbers if we want strict UOM precision,
    # or we can keep it for general cleanup but precision logic above is specific.
    # response = round_all_numbers(response)
    return response


# ==================== PATCH ENDPOINT - APPROVE BRANCHWISE ITEM ====================
@router.patch("/{itemCode}/approve")
async def approve_branchwise_item(httprequest:Request,
    itemCode: str,
    locationId: str = Query(..., description="Location ID"),
    request: ApproveItemRequest = Body(...),
    user = Depends(validate_token),
    permissions: dict = Depends(check_permission("yenerp","physicalstockvariancemodification","approve")),
):
    tenant_id = httprequest.state.tenant_id
    inventory_col = inventory_stock_collection(tenant_id)
    variance_col = approvedstocks_fg_collection(tenant_id)

    # Fetch UOM Map for precision
    uom_precision_map = await get_uom_precision_map()

    now = ist_now()
    item_code = itemCode.strip()
    location_id = locationId.strip()

    # 1️⃣ Validate item exists in master collection
    item_doc = await master_collection().find_one({"itemCode": item_code})
    if not item_doc:
        raise HTTPException(
            status_code=404,
            detail=f"ItemCode {item_code} Not Available in Branchwise Items collection",
        )

    # Get Precision
    uom_id = item_doc.get("item_Uom")
    precision = uom_precision_map.get(uom_id, 0)

    # 2️⃣ Fetch inventory record
    inventory_doc = await inventory_col.find_one(
        {"itemCode": item_code, "locationId": location_id}
    )

    if not inventory_doc:
        raise HTTPException(
            status_code=404,
            detail=f"No inventory record found for itemCode={item_code} and locationId={location_id}",
        )

    status = inventory_doc.get("status", "pending")

    # 3️⃣ Prevent double approval
    if status.lower() == "approved":
        return {
            "message": "Already approved. Approval not required.",
            "itemCode": item_code,
            "locationId": location_id,
            "status": status,
            "updatedAt": inventory_doc.get("updatedAt"),
        }

    # 4️⃣ Get values ONLY from DB
    system_stock = safe_float(inventory_doc.get("systemStock", 0))
    physical_stock = safe_float(inventory_doc.get("physicalStock", 0))
    actual_variance = safe_float(inventory_doc.get("variance", 0))

    # 5️⃣ If no variance → stop
    if actual_variance == 0:
        return {
            "message": "No variance found. Approval not required.",
            "itemCode": item_code,
            "locationId": location_id,
            "systemStock": round_by_precision(system_stock, precision),
            "physicalStock": round_by_precision(physical_stock, precision),
            "variance": 0,
            "status": "approved",
        }

    # 6️⃣ Update system stock using DB variance
    updated_system_stock = system_stock + actual_variance
    updated_system_stock = round_by_precision(updated_system_stock, precision)

    # 7️⃣ Update inventory collection
    await inventory_col.update_one(
        {"itemCode": item_code, "locationId": location_id},
        {
            "$set": {
                "previousSystemStock": round_by_precision(system_stock, precision),
                "systemStock": updated_system_stock,
                "physicalStock": round_by_precision(physical_stock, precision),
                "variance": 0,
                "status": "approved",
                "updatedAt": now,
                "lastUpdatedBy": request.approved_by,
            }
        },
    )

    # 8️⃣ Prepare variance history document
    variance_doc = {
        "itemCode": item_code,
        "itemName": item_doc.get("varianceName", ""),
        "locationId": location_id,
        "systemStockBefore": round_by_precision(system_stock, precision),
        "physicalClosing": round_by_precision(physical_stock, precision),
        "actualVariance": round_by_precision(actual_variance, precision),
        "systemStockAfter": updated_system_stock,
        "status": "approved",
        "approvedBy": request.approved_by,
        "description": request.description,
        "approvedAt": now,
    }

    # 9️⃣ INSERT INTO VARIANCE COLLECTION
    insert_result = await variance_col.insert_one(variance_doc)

    variance_doc["_id"] = str(insert_result.inserted_id)

    # 🔟 Final response
    return {
        "message": "Stock approved successfully",
        "data": variance_doc,
        "updatedSystemStock": updated_system_stock,
        "varianceAfterApproval": 0,
        "status": "approved",
    }


@router.get("/approved")
async def get_approved_items(request:Request,
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1, le=100),
    date: Optional[str] = None,  # expects "dd mm yyyy"
    user = Depends(validate_token),
    permissions: dict = Depends(check_permission("yenerp","physicalstockvariancemodification","read")),
):
    tenant_id = request.state.tenant_id
    collection = approvedstocks_fg_collection(tenant_id)
    # ... (Keep existing implementation, potentially add precision lookup if needed for display)
    # For brevity, assuming existing implementation is fine unless specific rounding is needed here too.
    # If needed, fetch uom_precision_map and round 'actualVariance', 'systemStockAfter' etc.

    query = {"status": "approved"}

    if date:
        try:
            dt = datetime.strptime(date, "%d %m %Y")
            next_day = dt + timedelta(days=1)
            query["approvedAt"] = {"$gte": dt, "$lt": next_day}
        except ValueError:
            pass

    skip = (page - 1) * limit

    projection = {
        "itemCode": 1,
        "itemName": 1,
        "locationId": 1,
        "systemStockBefore": 1,
        "physicalClosing": 1,
        "systemStockAfter": 1,
        "actualVariance": 1,
        "approvedBy": 1,
        "approvedAt": 1,
        "description": 1,
    }

    total, items = await asyncio.gather(
        collection.count_documents(query),
        collection
        .find(query, projection)
        .skip(skip)
        .limit(limit)
        .to_list(length=limit),
    )

    # Optional: Round the output based on UOM
    uom_map = await get_uom_precision_map()
    master_items = (
        await master_collection().find({}, {"itemCode": 1, "item_Uom": 1}).to_list(None)
    )
    item_uom_lookup = {i["itemCode"]: i.get("item_Uom") for i in master_items}

    for item in items:
        item["_id"] = str(item["_id"])
        if "approvedAt" in item and isinstance(item["approvedAt"], datetime):
            item["approvedAt"] = item["approvedAt"].strftime("%Y-%m-%d %H:%M:%S")

        # Apply rounding
        u_id = item_uom_lookup.get(item.get("itemCode"))
        prec = uom_map.get(u_id, 0)

        if "systemStockBefore" in item:
            item["systemStockBefore"] = round_by_precision(
                item["systemStockBefore"], prec
            )
        if "physicalClosing" in item:
            item["physicalClosing"] = round_by_precision(item["physicalClosing"], prec)
        if "systemStockAfter" in item:
            item["systemStockAfter"] = round_by_precision(
                item["systemStockAfter"], prec
            )
        if "actualVariance" in item:
            item["actualVariance"] = round_by_precision(item["actualVariance"], prec)

    return {"data": items, "page": page, "limit": limit, "total": total}


@router.post("/create")
async def create_inventory_snapshot_all_branches(request:Request, createdBy: str = "Inventory"):
    tenant_id = request.state.tenant_id
    inventory_col = closingstocks_collection(tenant_id)
    branch_col = location_collection()
    item_col = master_collection()
    stock_col = inventory_stock_collection(tenant_id)

    # Get UOM Map
    uom_precision_map = await get_uom_precision_map()

    now = ist_now()
    today = now.strftime("%Y-%m-%d")

    branches = await branch_col.find({"status": "active"}).to_list(None)
    if not branches:
        raise HTTPException(status_code=404, detail="No active branches")

    items = await item_col.find(
        {"status": "Active"}, {"itemCode": 1, "item_Uom": 1}
    ).to_list(None)
    if not items:
        raise HTTPException(status_code=404, detail="No active items")

    # Map itemCode -> uomId
    item_uom_lookup = {i["itemCode"]: i.get("item_Uom") for i in items}
    item_codes = list(item_uom_lookup.keys())

    created_count = 0
    skipped_count = 0
    total_inserted_items = 0

    for branch in branches:
        location_id = branch.get("locationId")
        if not location_id:
            continue

        exists = await inventory_col.find_one(
            {"locationId": location_id, "date": today}
        )
        if exists:
            skipped_count += 1
            continue

        stock_records = await stock_col.find(
            {"locationId": location_id},
            {
                "_id": 0,
                "itemCode": 1,
                "systemStock": 1,
                "physicalStock": 1,
                "variance": 1,
                "status": 1,
            },
        ).to_list(None)

        stock_map = {s["itemCode"]: s for s in stock_records if "itemCode" in s}

        docs = []
        for code in item_codes:
            stock = stock_map.get(code, {})

            # Get precision
            uom_id = item_uom_lookup.get(code)
            precision = uom_precision_map.get(uom_id, 0)

            system_stock = safe_float(stock.get("systemStock", 0.0))
            physical_stock = safe_float(stock.get("physicalStock", 0.0))
            variance = safe_float(stock.get("variance", 0.0))

            docs.append(
                {
                    "locationId": location_id,
                    "itemCode": code,
                    "systemStock": round_by_precision(system_stock, precision),
                    "physicalStock": round_by_precision(physical_stock, precision),
                    "variance": round_by_precision(variance, precision),
                    "status": stock.get("status", "pending"),
                    "date": today,
                    "createdBy": createdBy,
                    "createdAt": now,
                }
            )

        if docs:
            await inventory_col.insert_many(docs)
            created_count += 1
            total_inserted_items += len(docs)

    response = {
        "message": "Daily inventory snapshot created (all branches, all items)",
        "createdBranches": created_count,
        "skippedBranches": skipped_count,
        "totalInsertedItems": total_inserted_items,
        "date": today,
    }
    return force_float_conversion(response)


@router.get("/items", response_model=List[ItemResponse])
async def get_item_names(request:Request,
    page: int = Query(1, ge=1),
    limit: int = Query(10, le=100),
    search: str | None = None,
):
    tenant_id = request.state.tenant_id
    # ... Existing implementation is fine
    collection = master_collection()

    query = {}
    if search:
        query = {"varianceName": {"$regex": search, "$options": "i"}}

    cursor = (
        collection.find(
            query,
            {
                "_id": 0,
                "itemCode": 1,
                "varianceName": 1,
            },
        )
        .skip((page - 1) * limit)
        .limit(limit)
    )

    items = []
    async for doc in cursor:
        items.append(
            {
                "itemCode": doc["itemCode"],
                "varianceName": doc["varianceName"],
            }
        )

    return items


# -------------------- STOCK LEDGER --------------------


async def get_closing_stock_for_date(tenant_id,location_id, item_code, day):
    coll = closingstocks_collection(tenant_id)

    doc = await coll.find_one(
        {
            "locationId": location_id,
            "itemCode": item_code,
            "date": day.isoformat(),
        },
        projection={"physicalStock": 1, "systemStock": 1, "variance": 1, "status": 1},
        sort=[("date", -1)],
    )

    if doc:
        stock = (
            doc.get("physicalStock")
            if doc.get("physicalStock") is not None
            else doc.get("systemStock")
        )
        return safe_float(stock)

    return None


async def get_item_details(item_code: str) -> dict:
    item = await master_collection().find_one(
        {"itemCode": item_code}, {"varianceName": 1, "item_Uom": 1}
    )

    variance_name = item.get("varianceName", item_code)
    uom_id = item.get("item_Uom")

    uom_str = ""
    precision = 0  # Default
    if uom_id:
        uom_doc = await uom_collection().find_one({"uomId": uom_id})
        if uom_doc:
            uom_str = uom_doc.get("uom", "")
            precision = uom_doc.get("precision", 0)

    return {
        "varianceName": variance_name,
        "uom": uom_str,
        "precision": precision,  # Return precision
    }


async def build_stock_ledger(tenant_id,location_id, item_code, from_date, to_date):
    ledger = []

    # Fetch details including precision
    item_details = await get_item_details(item_code)
    variance_name = item_details["varianceName"]
    uom_str = item_details["uom"]
    precision = item_details["precision"]

    running_stock = 0
    opening_balance = None

    for current_date in daterange(from_date, to_date):
        dispatch = await get_dispatch_mapping_agg(location_id, current_date)
        sales = await get_sales_mapping_agg(location_id, current_date)
        sales_return = await get_sales_return_mapping_agg(location_id, current_date)
        warehouse = await get_warehousereturn_mapping_agg(location_id, current_date)
        wastage = await get_wastagereturn_mapping_agg(location_id, current_date)
        transfer = await get_stock_transfer_mapping_agg(location_id, current_date)

        # Extract Quantities
        dispatched_qty = dispatch.get(item_code, {}).get("dispatchedQty", 0)
        sales_qty = sales["full_map"].get(item_code, {}).get("salesQty", 0)
        sales_return_qty = (
            sales_return["full_map"].get(item_code, {}).get("returnQty", 0)
        )
        warehouse_qty = (
            warehouse["full_map"].get(item_code, {}).get("wastageReceivedQty", 0)
        )
        wastage_qty = wastage["full_map"].get(item_code, {}).get("wastageReturnQty", 0)

        transfer_map = transfer["full_map"]
        transfer_in = transfer_map.get(item_code, {}).get("transferInQty", 0)
        transfer_out = transfer_map.get(item_code, {}).get("transferOutQty", 0)

        # Calculate Closing
        db_closing = await get_closing_stock_for_date(
            tenant_id,location_id, item_code, current_date
        )
        if db_closing is not None:
            closing_stock = db_closing
        else:
            closing_stock = (
                running_stock
                + dispatched_qty
                + transfer_in
                - transfer_out
                - sales_qty
                + sales_return_qty
                - warehouse_qty
                - wastage_qty
            )

        if opening_balance is None:
            opening_balance = running_stock

        # Apply Precision to Ledger Row
        ledger.append(
            {
                "date": current_date.isoformat(),
                "openingStock": round_by_precision(running_stock, precision),
                "dispatchQty": round_by_precision(dispatched_qty, precision),
                "salesQty": round_by_precision(sales_qty, precision),
                "salesReturnQty": round_by_precision(sales_return_qty, precision),
                "stockTransferInQty": round_by_precision(transfer_in, precision),
                "stockTransferOutQty": round_by_precision(transfer_out, precision),
                "wastageReceivedQty": round_by_precision(warehouse_qty, precision),
                "wastageReturnQty": round_by_precision(wastage_qty, precision),
                "closingStock": round_by_precision(closing_stock, precision),
            }
        )

        running_stock = closing_stock

    return {
        "varianceName": variance_name,
        "uom": uom_str,
        "openingBalance": round_by_precision(opening_balance or 0, precision),
        "openingDate": from_date.isoformat(),
        "closingBalance": round_by_precision(running_stock, precision),
        "transactions": ledger,
    }


@router.get("/stock-ledger/transactions")
async def stock_ledger_transactions(request:Request,
    locationId: str = Query(...),
    itemCode: List[str] = Query(...),
    fromDate: date = Query(...),
    toDate: date = Query(...),
    user = Depends(validate_token),
    permissions: dict = Depends(check_permission("yenerp","stockledger","read")), # ✅ ADD
):
    tenant_id = request.state.tenant_id
    result = []

    for code in itemCode:
        ledger = await build_stock_ledger(
            tenant_id,location_id=locationId, item_code=code, from_date=fromDate, to_date=toDate
        )
        result.append(
            {
                "itemCode": code,
                "varianceName": ledger["varianceName"],
                "uom": ledger["uom"],
                "openingDate": ledger["openingDate"],
                "openingBalance": ledger["openingBalance"],
                "closingBalance": ledger["closingBalance"],
                "transactions": ledger["transactions"],
            }
        )

    return result


@router.get("/stock-ledger/transactions/excel")
async def export_stock_ledger_excel(request:Request,
    locationName: str = Query(...),
    itemCodes: List[str] = Query(...),
    fromDate: date = Query(...),
    toDate: date = Query(...),
    
    user = Depends(validate_token),
    permissions: dict = Depends(check_permission("yenerp","stockledger","read")), # ✅ ADD
):
    tenant_id = request.state.tenant_id
    wb = Workbook()
    wb.remove(wb.active)

    for code in itemCodes:
        ledger = await build_stock_ledger(
            location_id=locationName,
            item_code=code,
            from_date=fromDate,
            to_date=toDate,
        )

        safe_title = ledger["varianceName"][:31]

        ws = wb.create_sheet(title=safe_title)

        ws.append([f"Variance Name: {ledger['varianceName']}"])
        ws.append([f"Location: {locationName}"])
        ws.append([f"Date Range: {fromDate} to {toDate}"])
        ws.append([f"UOM: {ledger.get('uom', '')}"])
        ws.append([])

        ws.append(
            [
                "Date",
                "Opening Stock",
                "Dispatch",
                "Sales",
                "Sales Return",
                "Stock Transfer In",
                "Stock Transfer Out",
                "Wastage",
                "Closing Stock",
                "UOM",
            ]
        )

        header_fill = PatternFill("solid", fgColor="4F81BD")
        header_font = Font(color="FFFFFF", bold=True)
        for col in range(1, 11):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.fill = header_fill
            cell.font = header_font

        uom = ledger.get("uom", "")

        for row in ledger["transactions"]:
            ws.append(
                [
                    row["date"],
                    row["openingStock"],
                    row["dispatchQty"],
                    row["salesQty"],
                    row["salesReturnQty"],
                    row["stockTransferInQty"],
                    row["stockTransferOutQty"],
                    row["wastageReceivedQty"] + row["wastageReturnQty"],
                    row["closingStock"],
                    uom,
                ]
            )

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"stock_ledger_{locationName}_{fromDate}_{toDate}.xlsx"

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
