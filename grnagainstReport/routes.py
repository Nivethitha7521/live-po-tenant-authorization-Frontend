import io
from math import ceil
from typing import List, Optional
from fastapi import APIRouter, HTTPException, Query
from datetime import datetime
from datetime import datetime, timezone
from fastapi import APIRouter, Query, HTTPException
from fastapi.responses import StreamingResponse
from datetime import datetime, timezone, time
from typing import Optional, List
import pandas as pd
from openpyxl.utils import get_column_letter
import io
from fastapi import Depends,Request
from dependencies.auth import validate_token
from middlewares.permission_middleware import check_permission
from fastapi.responses import StreamingResponse
import pandas as pd


from ApInvoiceReport.models import DropdownResponse
from db.collections import (
    grn_collection as get_grn_collection,
)
from .models import (
    PaginatedReportResponse,
)


router = APIRouter()




# Helper function to safely get value from array
def safe_index(field_list, idx):
    if isinstance(field_list, list) and len(field_list) > idx:
        return field_list[idx]
    return None


def format_datetime(dt):
    """Safely format a datetime object or ISO string."""
    if not dt:
        return None, None
    if isinstance(dt, str):
        # convert ISO string to datetime
        try:
            dt = datetime.fromisoformat(dt.replace("Z", "+00:00"))
        except Exception:
            return None, None
    return dt.strftime("%d-%m-%Y"), dt.strftime("%H:%M")




@router.get("/date-dropdown", response_model=DropdownResponse)
async def get_dispatch_date_dropdown(request: Request, user=Depends(validate_token),
    permissions=Depends(check_permission("yenerp", "purchaseorderreport", "read"))):
    tenant_id = request.state.tenant_id   # 🔥 ADD THIS
    collection = get_grn_collection(tenant_id) 

    pipeline = [
        {"$match": {"createdDate": {"$type": "date"}}},
        {
            "$group": {
                "_id": None,
                "years": {"$addToSet": {"$year": "$createdDate"}},
                "months": {"$addToSet": {"$month": "$createdDate"}},
                "days": {"$addToSet": {"$dayOfMonth": "$createdDate"}},
            }
        },
    ]

    result = await collection.aggregate(pipeline).to_list(1)

    if not result:
        return DropdownResponse(yearIn=[], monthIn=[], daysIn=[])

    return DropdownResponse(
        yearIn=sorted(map(str, result[0]["years"])),
        monthIn=sorted(f"{m:02d}" for m in result[0]["months"]),
        daysIn=sorted(result[0]["days"]),
    )


@router.get(
    "/report", response_model=PaginatedReportResponse, summary="GRN against Report"
)
async def get_grn_against_unique_grn(request:Request,
    page: int = Query(1, ge=1),
    limit: Optional[int] = Query(30, ge=1),
    startDate: Optional[datetime] = Query(None),
    endDate: Optional[datetime] = Query(None),
    invoiceNo: Optional[List[str]] = Query(None),
    vendorName: Optional[List[str]] = Query(None),
    status: Optional[List[str]] = Query(None),
    user=Depends(validate_token),   # 🔥 ADD
    permissions=Depends(check_permission("yenerp", "purchaseorderreport", "read"))
):
    tenant_id = request.state.tenant_id   # 🔥 ADD THIS
    collection = get_grn_collection(tenant_id)
    try:
        pipeline = []

        base_filter = {}
        # 1. Base filter: only APInvoiceConverted
        if status:
            base_filter["status"] = {"$in": status}

        if invoiceNo:
            base_filter["invoiceNo"] = {"$in": invoiceNo}
        if vendorName:
            base_filter["vendorName"] = {"$in": vendorName}
        pipeline.append({"$match": base_filter})

        # 2. Lookup AP Invoices
        pipeline.append(
            {
                "$lookup": {
                    "from": "apInvoice",
                    "let": {"grn_id": "$grnId", "grn_randomId": "$randomId"},
                    "pipeline": [
                        {
                            "$match": {
                                "$expr": {
                                    "$or": [
                                        {"$eq": ["$grnId", "$$grn_id"]},
                                        {"$eq": ["$grnRandomId", "$$grn_randomId"]},
                                    ]
                                }
                            }
                        }
                    ],
                    "as": "apInvoiceData",
                }
            }
        )

        pipeline.append(
            {"$unwind": {"path": "$apInvoiceData", "preserveNullAndEmptyArrays": True}}
        )

        # 3. Date filter
        if startDate or endDate:
            if not startDate:
                startDate = endDate
            if not endDate:
                endDate = startDate

            start_dt = datetime.combine(startDate.date(), datetime.min.time()).replace(
                tzinfo=timezone.utc
            )
            end_dt = datetime.combine(endDate.date(), datetime.max.time()).replace(
                tzinfo=timezone.utc
            )

            pipeline.append(
                {
                    "$match": {
                        "$or": [
                            {"createdDate": {"$gte": start_dt, "$lte": end_dt}},
                            {
                                "apInvoiceData.createdDate": {
                                    "$gte": start_dt,
                                    "$lte": end_dt,
                                }
                            },
                        ]
                    }
                }
            )

        # 4. Add computed totals
        pipeline.append(
            {
                "$addFields": {
                    "netAmount": {
                        "$sum": {
                            "$map": {
                                "input": {"$ifNull": ["$itemDetails", []]},
                                "as": "i",
                                "in": {"$ifNull": ["$$i.totalPrice", 0]},
                            }
                        }
                    },
                    "totalBefTaxDiscountAmount": {
                        "$sum": {
                            "$map": {
                                "input": {"$ifNull": ["$itemDetails", []]},
                                "as": "i",
                                "in": {"$ifNull": ["$$i.befTaxDiscountAmount", 0]},
                            }
                        }
                    },
                    "totalTaxAmount": {
                        "$sum": {
                            "$map": {
                                "input": {"$ifNull": ["$itemDetails", []]},
                                "as": "i",
                                "in": {"$ifNull": ["$$i.taxAmount", 0]},
                            }
                        }
                    },
                    "totalFinalPrice": {
                        "$sum": {
                            "$map": {
                                "input": {"$ifNull": ["$itemDetails", []]},
                                "as": "i",
                                "in": {"$ifNull": ["$$i.finalPrice", 0]},
                            }
                        }
                    },
                }
            }
        )

        # 5. Group by GRN
        pipeline.append(
            {
                "$group": {
                    "_id": "$randomId",
                    "grnId": {"$first": "$grnId"},
                    "grnDate": {"$first": "$grnDate"},
                    "apcreatedDate": {"$first": "$apInvoiceData.createdDate"},
                    "invoiceNo": {"$first": "$invoiceNo"},
                    "createdDate": {"$first": "$createdDate"},
                    "vendorName": {"$first": "$vendorName"},
                    "randomId": {"$first": "$randomId"},
                    "befTaxDiscountAmount": {"$first": "$totalBefTaxDiscountAmount"},
                    "taxAmount": {"$first": "$totalTaxAmount"},
                    "finalPrice": {"$first": "$totalFinalPrice"},
                    "apRandomId": {"$first": "$apInvoiceData.randomId"},
                    "apRemarks": {"$first": "$apInvoiceData.comments"},
                    "netAmount": {"$first": "$netAmount"},
                    "apInvoiceDate": {"$first": "$apInvoiceData.apinvoiceDate"},
                    "status": {"$first": "$status"},
                }
            }
        )

        # 6. Project final fields
        pipeline.append(
            {
                "$project": {
                    "_id": 0,
                    "grnId": 1,
                    "grnDate": {
                        "$dateToString": {"format": "%Y-%m-%d", "date": "$grnDate"}
                    },
                    "apcreatedDate": {
                        "$dateToString": {
                            "format": "%Y-%m-%d",
                            "date": "$apcreatedDate",
                        }
                    },
                    "invoiceNo": 1,
                    "createdDate": {
                        "$dateToString": {
                            "format": "%Y-%m-%d",
                            "date": "$createdDate",
                        }
                    },
                    "vendorName": 1,
                    "randomId": 1,
                    "befTaxDiscountAmount": 1,
                    "taxAmount": 1,
                    "finalPrice": 1,
                    "apRandomId": 1,
                    "apRemarks": 1,
                    "netAmount": 1,
                    "apInvoiceDate": 1,
                    "status": 1,
                }
            }
        )

        # 7. Sort by AP creation date descending
        pipeline.append({"$sort": {"apcreatedDate": -1, "createdDate": -1}})

        # 8. Pagination
        if limit:
            pipeline.append({"$skip": (page - 1) * limit})
            pipeline.append({"$limit": limit})

        # Execute pipeline
        cursor = collection.aggregate(pipeline)
        results = await cursor.to_list(length=None)
        total_records = await collection.count_documents(base_filter)
        total_pages = (total_records + (limit or 30) - 1) // (limit or 30)

        return PaginatedReportResponse(
            page=page,
            limit=limit,
            totalRecords=total_records,
            totalPages=total_pages,
            items=results,
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def fmt_date(dt):
    try:
        if isinstance(dt, datetime):
            return dt.strftime("%m-%d-%Y")  # MM-DD-YYYY
        return dt or ""
    except Exception as e:
        return ""


@router.get("/export", summary="Export Grnagainst Report Excel")
async def export_grn_against_excel(request:Request,
    startDate: Optional[datetime] = Query(None),
    endDate: Optional[datetime] = Query(None),
    invoiceNo: Optional[List[str]] = Query(None),
    vendorName: Optional[List[str]] = Query(None),
    status: Optional[List[str]] = Query(None),
    user=Depends(validate_token),
    permissions=Depends(check_permission("yenerp", "purchaseorderreport", "read"))
):
    tenant_id = request.state.tenant_id  
    collection = get_grn_collection(tenant_id) 
    try:
        pipeline = []
        base_filter = {}
        if status:
            base_filter["status"] = {"$in": status}
        if invoiceNo:
            base_filter["invoiceNo"] = {"$in": invoiceNo}
        if vendorName:
            base_filter["vendorName"] = {"$in": vendorName}
        if base_filter:
            pipeline.append({"$match": base_filter})

        # Lookup AP Invoices
        pipeline.append(
            {
                "$lookup": {
                    "from": "apInvoice",
                    "let": {"grn_id": "$grnId", "grn_randomId": "$randomId"},
                    "pipeline": [
                        {
                            "$match": {
                                "$expr": {
                                    "$or": [
                                        {"$eq": ["$grnId", "$$grn_id"]},
                                        {"$eq": ["$grnRandomId", "$$grn_randomId"]},
                                    ]
                                }
                            }
                        }
                    ],
                    "as": "apInvoiceData",
                }
            }
        )

        pipeline.append(
            {"$unwind": {"path": "$apInvoiceData", "preserveNullAndEmptyArrays": True}}
        )

        # Date filter
        if startDate or endDate:
            if not startDate:
                startDate = endDate
            if not endDate:
                endDate = startDate
            start_dt = datetime.combine(startDate.date(), time.min).replace(
                tzinfo=timezone.utc
            )
            end_dt = datetime.combine(endDate.date(), time.max).replace(
                tzinfo=timezone.utc
            )
            pipeline.append(
                {
                    "$match": {
                        "$or": [
                            {"createdDate": {"$gte": start_dt, "$lte": end_dt}},
                            {
                                "apInvoiceData.createdDate": {
                                    "$gte": start_dt,
                                    "$lte": end_dt,
                                }
                            },
                        ]
                    }
                }
            )

        # Add computed totals
        pipeline.append(
            {
                "$addFields": {
                    "netAmount": {
                        "$sum": {
                            "$map": {
                                "input": {"$ifNull": ["$itemDetails", []]},
                                "as": "i",
                                "in": {"$ifNull": ["$$i.totalPrice", 0]},
                            }
                        }
                    },
                    "totalTaxAmount": {
                        "$sum": {
                            "$map": {
                                "input": {"$ifNull": ["$itemDetails", []]},
                                "as": "i",
                                "in": {"$ifNull": ["$$i.taxAmount", 0]},
                            }
                        }
                    },
                    "totalFinalPrice": {
                        "$sum": {
                            "$map": {
                                "input": {"$ifNull": ["$itemDetails", []]},
                                "as": "i",
                                "in": {"$ifNull": ["$$i.finalPrice", 0]},
                            }
                        }
                    },
                    "item_service": {
                        "$cond": {
                            "if": {
                                "$gt": [{"$size": {"$ifNull": ["$itemDetails", []]}}, 0]
                            },
                            "then": "I",
                            "else": "S",
                        }
                    },
                    "gst_bos": {
                        "$cond": {
                            "if": {
                                "$gt": [
                                    {
                                        "$strLenCP": {
                                            "$ifNull": [{"$toString": "$gstNumber"}, ""]
                                        }
                                    },
                                    0,
                                ]
                            },
                            "then": "GST",
                            "else": "BOS",
                        }
                    },
                }
            }
        )

        # Group by GRN
        pipeline.append(
            {
                "$group": {
                    "_id": "$randomId",
                    "GRPO_NO": {"$first": "$randomId"},
                    "GRPO_DATE": {"$first": "$grnDate"},
                    "A/P_INVOICE_DATE": {"$first": "$apInvoiceData.apinvoiceDate"},
                    "A/P_INVOICE_NO": {"$first": "$apInvoiceData.randomId"},
                    "GST/BOS": {"$first": "$gst_bos"},
                    "ITEM/SERVICE": {"$first": "$item_service"},
                    "USER NAME": {"$first": "$userName"},
                    "VENDOR_REF_NO": {"$first": "$invoiceNo"},
                    "VENDOR_NAME": {"$first": "$vendorName"},
                    "Status": {"$first": "$status"},
                    "Bill to": {"$first": "$billingAddress"},
                    "GST_NO": {"$first": "$gstNumber"},
                    "NET_AMOUNT": {"$first": "$netAmount"},
                    "TAX_AMOUNT": {"$first": "$totalTaxAmount"},
                    "GROSS_AMOUNT": {"$first": "$totalFinalPrice"},
                    "GRPO Remarks": {"$first": "$comments"},
                    "A/P Remarks": {"$first": "$apInvoiceData.comments"},
                    "apcreatedDate": {"$first": "$apInvoiceData.createdDate"},
                    "createdDate": {"$first": "$createdDate"},
                }
            }
        )

        # Project final fields
        pipeline.append(
            {
                "$project": {
                    "_id": 0,
                    "GRPO_NO": 1,
                    "GRPO_DATE": 1,
                    "A/P_INVOICE_DATE": 1,
                    "A/P_INVOICE_NO": 1,
                    "GST/BOS": 1,
                    "ITEM/SERVICE": 1,
                    "USER NAME": 1,
                    "VENDOR_REF_NO": 1,
                    "VENDOR_NAME": 1,
                    "Status": 1,
                    "Bill to": 1,
                    "GST_NO": 1,
                    "NET_AMOUNT": 1,
                    "TAX_AMOUNT": 1,
                    "GROSS_AMOUNT": 1,
                    "GRPO Remarks": 1,
                    "A/P Remarks": 1,
                }
            }
        )

        # Sort
        pipeline.append({"$sort": {"apcreatedDate": -1, "createdDate": -1}})
        cursor = collection.aggregate(pipeline)
        results = await cursor.to_list(length=None)
        if not results:
            raise HTTPException(status_code=404, detail="No data found for export")

        # Convert to DataFrame
        df = pd.DataFrame(results)

        # Force IDs as text for Excel
        for col in ["VENDOR_REF_NO", "GRPO_NO", "A/P_INVOICE_NO"]:
            if col in df.columns:
                df[col] = df[col].apply(lambda x: str(x) if x is not None else "")

        # Format date columns
        for col in ["GRPO_DATE", "A/P_INVOICE_DATE"]:
            if col in df.columns:
                df[col] = df[col].apply(lambda x: fmt_date(x) if pd.notna(x) else "")

        # Format numeric columns
        for col in ["NET_AMOUNT", "TAX_AMOUNT", "GROSS_AMOUNT"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

        # Create Excel in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Sheet1", index=False)
            worksheet = writer.sheets["Sheet1"]

            # Force Excel text format for ID columns
            from openpyxl.styles import numbers

            for col_idx, col_name in enumerate(df.columns, 1):
                if col_name in ["VENDOR_REF_NO", "GRPO_NO", "A/P_INVOICE_NO"]:
                    for row in range(2, len(df) + 2):
                        worksheet.cell(row=row, column=col_idx).number_format = (
                            numbers.FORMAT_TEXT
                        )

                # Auto-adjust column width
                max_length = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in worksheet[get_column_letter(col_idx)]
                )
                worksheet.column_dimensions[get_column_letter(col_idx)].width = (
                    max_length + 2
                )

        output.seek(0)
        download_time = datetime.now().strftime("%d-%m-%Y_%H-%M")
        filename = f"GRNagainst_YenERP_{download_time}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")
